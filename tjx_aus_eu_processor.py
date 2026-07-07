import customtkinter as ctk
from tkinterdnd2 import TkinterDnD, DND_FILES
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy as copy_style
from shutil import copy2
import threading
import os
import traceback
from tkinter import filedialog, messagebox
import re
import queue
import subprocess
import sys
import time
import csv
import tempfile
import json
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

# -----------------------
# Appearance / Theme
# -----------------------
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# Soft modal tones
BG_MAIN = "#F3F4F6"
CARD_BG = "#FFFFFF"
TEXT_MUTED = "#6B7280"
TEXT_DARK = "#111827"
BORDER_LIGHT = "#E5E7EB"
PRIMARY = "#2563EB"
PRIMARY_HOVER = "#EEF2FF"
SUCCESS = "#15803D"
WARNING = "#B45309"
ERROR = "#B91C1C"
APP_NAME = "Teq Shipment Docs Generator"
OLD_APP_NAME = "TJX AUS EU PL CI Processor"
SETTINGS_FILE = Path(os.getenv("APPDATA") or Path.home()) / APP_NAME / "settings.json"
OLD_SETTINGS_FILE = Path(os.getenv("APPDATA") or Path.home()) / OLD_APP_NAME / "settings.json"
ACTIVE_PROCESS_LOCK = threading.Lock()
ACTIVE_PROCESSES = set()


class ProcessingCancelled(Exception):
    pass


def _check_cancel(cancel_event=None):
    if cancel_event is not None and cancel_event.is_set():
        raise ProcessingCancelled("Cancelled by user")


def _terminate_active_processes():
    with ACTIVE_PROCESS_LOCK:
        processes = list(ACTIVE_PROCESSES)
    for process in processes:
        if process.poll() is not None:
            continue
        try:
            if os.name == "nt":
                subprocess.run(
                    ["taskkill", "/PID", str(process.pid), "/T", "/F"],
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    creationflags=subprocess.CREATE_NO_WINDOW,
                    check=False,
                )
            else:
                process.terminate()
        except Exception:
            pass

# -----------------------
# Core Logic (kept)
# -----------------------
def extract_po_number(sonum):
    """Extract only the 8 digits after 'CA'."""
    match = re.search(r"CA(\d{8})", str(sonum))
    return match.group(1) if match else str(sonum)


def get_ship_date_filename_token(input_file):
    """
    Read ship date from sheet "Master Cartons Qty" and format as YYYYMMDD.
    Returns None if sheet/column/value is unavailable.
    """
    try:
        df_ship = pd.read_excel(input_file, sheet_name="Master Cartons Qty")
    except Exception:
        return None

    if df_ship.empty:
        return None

    columns = list(df_ship.columns)
    preferred = ["Ship Date", "ShipDate", "Ship_Date", "Ship Dt", "Date Shipped"]
    candidates = [c for c in preferred if c in columns]

    if not candidates:
        for c in columns:
            c_text = str(c).strip().lower()
            if "ship" in c_text and "date" in c_text:
                candidates.append(c)

    for col in candidates:
        series = df_ship[col].dropna()
        if series.empty:
            continue

        parsed = pd.to_datetime(series, errors="coerce")
        parsed = parsed.dropna()
        if not parsed.empty:
            return parsed.iloc[0].strftime("%Y%m%d")

        raw_text = str(series.iloc[0]).strip()
        if raw_text:
            safe = re.sub(r"[^A-Za-z0-9_-]", "-", raw_text)
            return safe[:40]

    return None


def _clean_excel_value(value):
    if pd.isna(value):
        return ""
    return value


def _clean_text(value):
    value = _clean_excel_value(value)
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _normalize_retailer_key(retailer):
    key = str(retailer).strip().upper()
    return "TJXCA" if key == "CA" else key


def _format_composition_text(value):
    text = _clean_text(value)
    if not text or text in ("0", "0.0"):
        return ""
    text = re.sub(r"[\r\n]+", ", ", text)
    text = re.sub(r"\s{2,}", ", ", text)
    text = re.sub(r"\s*,\s*", ", ", text)
    text = re.sub(r"(,\s*)+", ", ", text)
    return text.strip(" ,")


def _format_ctn_dimensions_cm(value):
    text = _clean_text(value)
    parts = [part for part in re.split(r"\s*[*xX]+\s*", text) if part]
    if len(parts) != 3:
        return text

    numbers = []
    for part in parts:
        number = _to_number(part, None)
        if number is None:
            return text
        numbers.append(number)

    convert_from_mm = any(abs(number) >= 100 for number in numbers)
    converted = []
    for number in numbers:
        if convert_from_mm:
            number = number / 10
        converted.append(str(int(number)) if float(number).is_integer() else str(round(number, 2)))
    return "*".join(converted)


def _carton_cbm_from_dimensions(value):
    text = _clean_text(value)
    parts = [part for part in re.split(r"\s*[*xX]+\s*", text) if part]
    if len(parts) != 3:
        return 0
    numbers = []
    for part in parts:
        number = _to_number(part, None)
        if number is None:
            return 0
        numbers.append(number)
    divisor = 1000
    cbm = 1
    for number in numbers:
        cbm *= number / divisor
    return round(cbm, 3)


def _clean_po(value):
    text = _clean_text(value)
    return text[:-2] if text.endswith(".0") else text


def _to_number(value, default=0):
    if pd.isna(value):
        return default
    converted = pd.to_numeric(value, errors="coerce")
    if pd.isna(converted):
        return default
    return float(converted)


def _to_int(value, default=0):
    return int(round(_to_number(value, default)))


def _as_excel_number(value):
    number = _to_number(value, None)
    if number is None:
        return ""
    return int(number) if float(number).is_integer() else round(float(number), 4)


def _safe_path_token(value, fallback="Shipment"):
    text = _clean_text(value)
    if not text:
        text = fallback
    parsed = pd.to_datetime(text, errors="coerce")
    if not pd.isna(parsed):
        text = parsed.strftime("%Y-%m-%d")
    text = re.sub(r"[<>:\"/\\|?*]+", "-", text)
    text = re.sub(r"\s+", " ", text).strip(" .-_")
    return text[:80] or fallback


def _shipment_key(row, cols, fallback):
    shipment_col = cols.get("Shipment")
    if shipment_col and shipment_col in row.index:
        shipment = _clean_text(row.get(shipment_col))
        if shipment:
            return _safe_path_token(shipment, fallback)
    return _safe_path_token(fallback, "Shipment")


def _resolve_template(template_name):
    candidates = [Path(template_name)]
    script_dir = Path(__file__).parent
    candidates.append(script_dir / template_name)
    candidates.append(script_dir / "templates" / template_name)

    if getattr(sys, "frozen", False):
        candidates.append(Path(sys._MEIPASS) / template_name)
        candidates.append(Path(sys.executable).parent / template_name)
        candidates.append(Path(sys.executable).parent / "_internal" / template_name)

    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def _strip_workbook_external_state(wb):
    if wb.defined_names:
        for name in list(wb.defined_names):
            del wb.defined_names[name]
    for ws in wb.worksheets:
        if ws.tables:
            for table_name in list(ws.tables.keys()):
                del ws.tables[table_name]


def _force_workbook_arial_12(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                font = copy_style(cell.font)
                font.name = "Arial"
                font.sz = 12
                cell.font = font


def _copy_row_format(ws, source_row, target_row, max_col=None):
    max_col = max_col or ws.max_column
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dest = ws.cell(target_row, col)
        if src.has_style:
            dest._style = copy_style(src._style)
        if src.number_format:
            dest.number_format = src.number_format
        if src.alignment:
            dest.alignment = copy_style(src.alignment)


def _copy_detail_rows_format(ws, source_row, start_row, end_row, max_col):
    if end_row < start_row:
        return
    for row in range(start_row, end_row + 1):
        _copy_row_format(ws, source_row, row, max_col=max_col)


def _find_row_by_value(ws, value, start_row=1, col=None):
    target = str(value).strip().lower()
    for row in range(start_row, ws.max_row + 1):
        cols = [col] if col else range(1, ws.max_column + 1)
        for c in cols:
            cell_value = ws.cell(row, c).value
            if str(cell_value).strip().lower() == target:
                return row
    return None


def _ensure_detail_capacity(ws, start_row, total_row, rows_needed, max_col, style_row=None):
    shifted_merged_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row and merged_range.max_row <= total_row:
            ws.unmerge_cells(str(merged_range))
        elif merged_range.min_row >= total_row:
            shifted_merged_ranges.append((
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            ))
            ws.unmerge_cells(str(merged_range))

    available = max(0, total_row - start_row)
    style_row = style_row or start_row
    if rows_needed > available:
        rows_to_add = rows_needed - available
        ws.insert_rows(total_row, rows_to_add)
        for min_row, min_col, max_row, max_col_range in shifted_merged_ranges:
            ws.merge_cells(
                start_row=min_row + rows_to_add,
                start_column=min_col,
                end_row=max_row + rows_to_add,
                end_column=max_col_range,
            )
        for row in range(total_row, total_row + rows_to_add):
            _copy_row_format(ws, style_row, row, max_col=max_col)
        total_row += rows_to_add
    else:
        for min_row, min_col, max_row, max_col_range in shifted_merged_ranges:
            ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col_range,
            )

    for row in range(start_row, total_row):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ != "MergedCell":
                cell.value = None
    return total_row


def _ensure_aus_ci_detail_capacity(ws, start_row, total_row, rows_needed, max_col):
    blank_row = total_row - 1
    shifted_merged_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row and merged_range.max_row <= blank_row:
            ws.unmerge_cells(str(merged_range))
        elif merged_range.min_row >= total_row:
            shifted_merged_ranges.append((
                merged_range.min_row,
                merged_range.min_col,
                merged_range.max_row,
                merged_range.max_col,
            ))
            ws.unmerge_cells(str(merged_range))

    available = max(0, blank_row - start_row)
    if rows_needed > available:
        rows_to_add = rows_needed - available
        ws.insert_rows(blank_row, rows_to_add)
        source_blank_row = blank_row + rows_to_add
        for row in range(blank_row, source_blank_row):
            _copy_row_format(ws, source_blank_row, row, max_col=max_col)
        blank_row += rows_to_add
        total_row += rows_to_add
        for min_row, min_col, max_row, max_col_range in shifted_merged_ranges:
            ws.merge_cells(
                start_row=min_row + rows_to_add,
                start_column=min_col,
                end_row=max_row + rows_to_add,
                end_column=max_col_range,
            )
    else:
        for min_row, min_col, max_row, max_col_range in shifted_merged_ranges:
            ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col_range,
            )

    for row in range(start_row, blank_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if type(cell).__name__ != "MergedCell":
                cell.value = None
    return total_row, blank_row


def _safe_div(numerator, denominator):
    denominator = _to_number(denominator)
    if not denominator:
        return 0
    return _to_number(numerator) / denominator


def _load_region_summary(input_file, region):
    try:
        df = pd.read_excel(input_file, sheet_name="Summary")
    except Exception:
        xls = pd.ExcelFile(input_file)
        if "Summary" not in xls.sheet_names:
            return None, "Missing required sheet: 'Summary'"
        df = pd.read_excel(input_file, sheet_name="Summary")

    column_map = {
        "PONum": "PONum",
        "SKU": "SKU",
        "ProductDescription": "ProductDescription",
        "Master_Qty": "Master Qty",
        "Carton_Qty": "Carton Qty",
        "Total_Units": "Total Units",
        "Price": "USD Price",
        "Amount": "Amount",
        "Factory": "Factory",
        "Teq_PO": "PO",
        "Weight_Un": "Weight Un",
        "Grand_Weight": "Grand Weight",
        "Grand_Net": "Grand Net",
        "Dimensions": "Dimentions",
        "Grand_Volume": "Grand Volume",
    }
    composition_columns = [
        "Material Composition",
        "Materail composition",
        "Material composition",
        "Composition",
    ]
    for composition_col in composition_columns:
        if composition_col in df.columns:
            column_map["Composition"] = composition_col
            break
    required = [
        column_map["PONum"],
        column_map["SKU"],
        column_map["ProductDescription"],
        column_map["Master_Qty"],
        column_map["Carton_Qty"],
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        return None, f"Missing required column(s) for {region}: {', '.join(missing)}"

    df = df.dropna(how="all").copy()
    df = df[df[column_map["PONum"]].notna() & df[column_map["SKU"]].notna()].copy()
    if df.empty:
        return None, f"No valid {region} summary rows found."

    for col in [
        column_map["Master_Qty"],
        column_map["Carton_Qty"],
        column_map["Total_Units"],
        column_map["Price"],
        column_map["Amount"],
        column_map["Weight_Un"],
        column_map["Grand_Weight"],
        column_map["Grand_Net"],
        column_map["Grand_Volume"],
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["_PONum_str"] = df[column_map["PONum"]].map(_clean_po)
    return (df, column_map), None


def _master_info_lookup(input_file):
    try:
        master = pd.read_excel(input_file, sheet_name="Master Info")
    except Exception:
        return {}

    if "SKU" not in master.columns:
        return {}

    lookup = {}
    for _, row in master.dropna(how="all").iterrows():
        sku = _clean_text(row.get("SKU"))
        if not sku or sku in lookup:
            continue
        material = ""
        for material_col in ["Material Composition", "Materail composition", "Material composition", "Composition"]:
            if material_col in master.columns:
                material = _format_composition_text(row.get(material_col))
                if material and material not in ("0", "0.0"):
                    break
        lookup[sku] = {
            "factory": _clean_text(row.get("Factory")),
            "dimensions": _clean_text(row.get("Dimentions")),
            "weight": _to_number(row.get("Weight")),
            "weight_un": _to_number(row.get("Weight Un")),
            "material": material,
            "dc_name": _clean_text(row.get("DC Name:")),
            "ship_to": _clean_text(row.get("Ship Merchandise To")),
        }
    return lookup


def _load_tjxca_summary(input_file):
    loaded, error = _load_region_summary(input_file, "TJXCA")
    if error:
        return None, error.replace("TJXCA", "TJX CA")
    return loaded, None


def _tjxca_master_lookups(input_file):
    try:
        master = pd.read_excel(input_file, sheet_name="Master Info")
    except Exception:
        return {}, {}, {}

    item_lookup = {}
    order_lookup = {}
    dc_lookup = {}
    for _, row in master.dropna(how="all").iterrows():
        sku = _clean_text(row.get("SKU"))
        if sku and sku not in item_lookup:
            item_lookup[sku] = {
                "factory": _clean_text(row.get("Factory")),
                "dimensions": _clean_text(row.get("Dimentions")),
                "weight": _to_number(row.get("Weight")),
                "weight_un": _to_number(row.get("Weight Un")),
            }

        dc = _clean_po(row.get("DC"))
        if dc:
            dc_lookup[dc] = {
                "dc": dc,
                "dc_name": _clean_text(row.get("DC Name:")) or "Winners Merchants International LP",
                "ship_to": _clean_text(row.get("Ship Merchandise To")),
            }

        order = _clean_po(row.get("Orders"))
        if order:
            order_lookup[order] = {
                "ticket": _clean_text(row.get("Ticket")) or "No",
                "inner": _clean_text(row.get("Inner")) or "No",
            }
    return item_lookup, order_lookup, dc_lookup


def _tjxca_destination_for_po(po, order_lookup, dc_lookup):
    po = _clean_po(po)
    destination = dict(order_lookup.get(po, {}))
    dc_info = dc_lookup.get(po[:2], {})
    destination.update({
        "dc": dc_info.get("dc", po[:2]),
        "dc_name": dc_info.get("dc_name", "Winners Merchants International LP"),
        "ship_to": dc_info.get("ship_to", ""),
        "ticket": destination.get("ticket", "No"),
        "inner": destination.get("inner", "No"),
    })
    return destination


def _row_composition(sku, master_lookup, row=None, cols=None):
    material = ""
    if row is not None and cols and cols.get("Composition") in row.index:
        material = _format_composition_text(row.get(cols["Composition"]))
    if not material or material in ("0", "0.0"):
        material = master_lookup.get(_clean_text(sku), {}).get("material", "")
    return material if material and material not in ("0", "0.0") else ""


def _build_ci_detail_rows(rows):
    ci_rows = []
    active_composition = None

    def flush_composition():
        nonlocal active_composition
        if active_composition:
            ci_rows.append({
                "row_type": "composition",
                "desc": f"({active_composition})",
                "composition": active_composition,
            })
        active_composition = None

    for item in rows:
        composition = item.get("composition", "")
        if composition != active_composition:
            flush_composition()
        ci_rows.append({**item, "row_type": "item"})
        active_composition = composition or None

    flush_composition()
    return ci_rows


def _ci_description(item):
    composition = item.get("composition", "")
    if composition:
        return f'{item["desc"]}\n({composition})'
    return item["desc"]


def _build_region_rows(po_df, cols, master_lookup):
    rows = []
    carton_cursor = 1
    for _, row in po_df.iterrows():
        cartons = _to_int(row[cols["Carton_Qty"]])
        if cartons <= 0:
            continue
        units = _to_int(row[cols["Master_Qty"]])
        total_units = _to_number(row.get(cols["Total_Units"], units * cartons))
        if not total_units:
            total_units = units * cartons
        end_carton = carton_cursor + cartons - 1
        ctn_range = f"{carton_cursor}-{end_carton}" if cartons > 1 else str(carton_cursor)
        sku = _clean_text(row[cols["SKU"]])
        master = master_lookup.get(sku, {})
        dimensions = _clean_text(row.get(cols["Dimensions"], "")) or master.get("dimensions", "")
        gross_total = _to_number(row.get(cols["Grand_Weight"], 0))
        if not gross_total and master.get("weight"):
            gross_total = _to_number(master["weight"]) * cartons
        net_total = _to_number(row.get(cols["Grand_Net"], 0))
        if not net_total:
            net_total = max(gross_total - cartons, 0) if gross_total else 0
        volume_total = _to_number(row.get(cols["Grand_Volume"], 0))
        desc = _clean_text(row[cols["ProductDescription"]])
        composition = _row_composition(sku, master_lookup, row=row, cols=cols)
        rows.append({
            "ctn": ctn_range,
            "po": _clean_po(row[cols["PONum"]]),
            "sku": sku,
            "desc": desc,
            "composition": composition,
            "units": units,
            "cartons": cartons,
            "total_units": _as_excel_number(total_units),
            "price": _as_excel_number(row.get(cols["Price"], 0)),
            "amount": _as_excel_number(row.get(cols["Amount"], _to_number(row.get(cols["Price"], 0)) * total_units)),
            "teq_po": _clean_text(row.get(cols["Teq_PO"], "")),
            "gross_each": _as_excel_number(_safe_div(gross_total, cartons)),
            "gross_total": _as_excel_number(gross_total),
            "net_each": _as_excel_number(_safe_div(net_total, cartons)),
            "net_total": _as_excel_number(net_total),
            "volume_total": _as_excel_number(volume_total),
            "dimensions": dimensions,
        })
        carton_cursor = end_carton + 1
    return rows


def _label_destination(region, po):
    region = str(region).strip().upper()
    if region == "AUS":
        return {
            "dept": "13-28",
            "country": "China",
            "vendor_packed": "No",
            "pre_ticketed": "No",
            "over_25kg": "No",
            "to": "TJX Australia Processing Centre\n25 Astoria Street Marsden Park, NSW 2765, Australia",
        }

    po_prefix = str(po)[:2]
    if po_prefix == "50":
        return {
            "dept": "36",
            "country": "China",
            "no": "No",
            "no2": "No",
            "vendor_packed": "No",
            "pre_ticketed": "No",
            "over_25kg": "No",
            "dc_name": "TJX UK",
            "ship_to": "73 Clarendon Road, Watford, Herts, UNITED KINGDOM, WD17 1TX",
        }
    if po_prefix == "55":
        return {
            "dept": "36",
            "country": "China",
            "no": "No",
            "no2": "No",
            "vendor_packed": "No",
            "pre_ticketed": "No",
            "over_25kg": "No",
            "dc_name": "TJX UK c/o TJX Distribution Ltd & Co KG",
            "ship_to": "Ben-Cammarata Strasse 1, Bergheim, GERMANY, 50126",
        }
    return None


def _build_label_rows(shipment_df, cols, master_lookup, region):
    region = str(region).strip().upper()
    label_rows = []
    for po in sorted(shipment_df["_PONum_str"].dropna().unique(), key=str):
        destination = _label_destination(region, po)
        if destination is None:
            return None, f"{region} PO '{po}' does not start with 50 or 55, so no label destination can be selected."

        po_df = shipment_df[shipment_df["_PONum_str"] == po].copy()
        total_cartons = _to_int(pd.to_numeric(po_df[cols["Carton_Qty"]], errors="coerce").fillna(0).sum())
        carton_number = 1

        for _, row in po_df.iterrows():
            cartons = _to_int(row[cols["Carton_Qty"]])
            if cartons <= 0:
                continue

            sku = _clean_text(row[cols["SKU"]])
            factory = _clean_text(row.get(cols["Factory"], "")) or master_lookup.get(sku, {}).get("factory", "")
            label_units = _as_excel_number(row[cols["Master_Qty"]])
            description = _clean_text(row[cols["ProductDescription"]])

            for _ in range(cartons):
                if region == "AUS":
                    label_rows.append([
                        po,
                        destination["dept"],
                        sku,
                        description,
                        label_units,
                        carton_number,
                        total_cartons,
                        destination["country"],
                        destination["vendor_packed"],
                        destination["pre_ticketed"],
                        destination["over_25kg"],
                        destination["to"],
                        factory,
                    ])
                else:
                    label_rows.append([
                        destination["dept"],
                        po,
                        sku,
                        description,
                        label_units,
                        carton_number,
                        total_cartons,
                        destination["country"],
                        destination["no"],
                        destination["no2"],
                        destination["vendor_packed"],
                        destination["pre_ticketed"],
                        destination["over_25kg"],
                        destination["dc_name"],
                        destination["ship_to"],
                        factory,
                    ])
                carton_number += 1

    return label_rows, None


def _write_labels_workbook(output_path, shipment_df, cols, master_lookup, region):
    region = str(region).strip().upper()
    if region == "AUS":
        headers = [
            "TJX PO", "Dept No", "Vendor Style ", "Description", "Total units",
            "Carton #", "Total Carton Qty", "Country ", "Vendor Packed",
            "Pre-Ticketed", ">25KG", "To", "Factory",
        ]
    else:
        headers = [
            "Dept No", "TJX PO", "Vendor Style", "Description", "Total Units",
            "Carton #", "Total Carton Qty", "Country", "No", "No2",
            "Vendor Packed", "Pre-Ticketed", ">25KG", "DC Name:",
            "Ship Merchandise To", "Factory",
        ]

    label_rows, error = _build_label_rows(shipment_df, cols, master_lookup, region)
    if error:
        return False, error
    if not label_rows:
        return False, f"No {region} label rows were generated."

    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"
    ws.append(headers)
    for row in label_rows:
        ws.append(row)

    last_row = len(label_rows) + 1
    last_col = len(headers)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName="LabelsTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.freeze_panes = "A2"

    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        1: 12, 2: 12, 3: 18, 4: 56, 5: 13, 6: 10, 7: 17, 8: 12,
        9: 12, 10: 13, 11: 12, 12: 46, 13: 18, 14: 34, 15: 58, 16: 18,
    }
    for col in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 16)
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _force_workbook_arial_12(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return True, None


def _build_tjxca_label_rows(shipment_df, cols, order_lookup, dc_lookup):
    label_rows = []
    for po in sorted(shipment_df["_PONum_str"].dropna().unique(), key=str):
        po_df = shipment_df[shipment_df["_PONum_str"] == po].copy()
        total_cartons = _to_int(pd.to_numeric(po_df[cols["Carton_Qty"]], errors="coerce").fillna(0).sum())
        carton_number = 1
        destination = _tjxca_destination_for_po(po, order_lookup, dc_lookup)

        for _, row in po_df.iterrows():
            cartons = _to_int(row[cols["Carton_Qty"]])
            if cartons <= 0:
                continue

            factory = _clean_text(row.get(cols["Factory"], ""))
            units = _as_excel_number(row[cols["Master_Qty"]])
            sku = _clean_text(row[cols["SKU"]])

            for _ in range(cartons):
                label_rows.append([
                    "36",
                    po,
                    sku,
                    _clean_text(row[cols["ProductDescription"]]),
                    units,
                    carton_number,
                    total_cartons,
                    "China",
                    "Yes",
                    destination.get("ticket", "No"),
                    "No",
                    destination.get("dc_name", "Winners Merchants International LP"),
                    destination.get("ship_to", ""),
                    destination.get("inner", "No"),
                    factory,
                ])
                carton_number += 1

    return label_rows


def _write_tjxca_labels_workbook(output_path, shipment_df, cols, order_lookup, dc_lookup):
    headers = [
        "Dept_No", "TJX_PO", "Vendor_Style", "Description", "Total_Units",
        "Carton_", "Total_Carton_Qty", "Country", "Vendor Packed",
        "PreTicketed", ">25KG", "DC Name:", "Ship_Merchandise_To",
        "Inner", "Factory",
    ]
    label_rows = _build_tjxca_label_rows(shipment_df, cols, order_lookup, dc_lookup)
    if not label_rows:
        return False, "No TJX CA label rows were generated."

    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"
    ws.append(headers)
    for row in label_rows:
        ws.append(row)

    last_row = len(label_rows) + 1
    last_col = len(headers)
    table = Table(displayName="LabelsTable", ref=f"A1:{get_column_letter(last_col)}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        1: 10, 2: 14, 3: 18, 4: 58, 5: 13, 6: 10, 7: 17, 8: 12,
        9: 14, 10: 13, 11: 12, 12: 34, 13: 62, 14: 10, 15: 18,
    }
    for col in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 16)
    for row in ws.iter_rows(min_row=2, max_row=last_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    _force_workbook_arial_12(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return True, None


def _label_template_name(region):
    region = str(region).strip().upper()
    if region == "AUS":
        return "TEMPLATE_LABELS_____________TJX AUS.docx"
    if region == "TJXCA":
        return "TJX CA FOB Labels template.docx"
    return "TEMPLATE_LABELS_____________TJX EU.docx"


def _merge_rows_by_factory(label_rows, region):
    region = str(region).strip().upper()
    rows_by_factory = {}
    for row in label_rows:
        if region == "TJXCA":
            factory = _clean_text(row[14]) or "Unknown Factory"
            ship_to = _clean_text(row[12])
            merge_row = {
                "Ship_Merchandise_To": ship_to,
                "TJX_PO": row[1],
                "Dept_No": row[0],
                "Vendor_Style": row[2],
                "Colour": "",
                "Total_Units": row[4],
                "Store_ready": row[8],
                "PreTicketed": row[9],
                "Inner": row[13],
                "Carton_": row[5],
                "Total_Carton_Qty": row[6],
                "COO": row[7],
                "Factory": factory,
            }
        elif region == "AUS":
            factory = _clean_text(row[12]) or "Unknown Factory"
            merge_row = {
                "TJX_PO": row[0],
                "Dept_No": row[1],
                "Vendor_Style_": row[2],
                "Description": row[3],
                "Total_units": row[4],
                "Carton_": row[5],
                "Total_Carton_Qty": row[6],
                "Country_": row[7],
                "Vendor_Packed": row[8],
                "PreTicketed": row[9],
                "M_25KG": row[10],
                "To": _clean_text(row[11]).replace("\n", ", "),
                "Factory": factory,
            }
        else:
            factory = _clean_text(row[15]) or "Unknown Factory"
            merge_row = {
                "TJX_PO": row[1],
                "Dept_No": row[0],
                "Vendor_Style_": row[2],
                "Description": row[3],
                "Total_units": row[4],
                "Carton_": row[5],
                "Total_Carton_Qty": row[6],
                "Country_": row[7],
                "Vendor_Packed": row[10],
                "PreTicketed": row[11],
                "M_25KG": row[12],
                "To": f"{row[13]}, {row[14]}".strip(" ,"),
                "Factory": factory,
            }
        rows_by_factory.setdefault(factory, []).append(merge_row)
    return rows_by_factory


def _write_mail_merge_pdf(template_path, csv_path, pdf_path, cancel_event=None):
    if os.name != "nt":
        return False, "Factory label PDF mail merge currently requires the Windows version of Microsoft Word. CI/PL workbooks and Indigo PDF labels can still be generated on macOS."

    ps_script = r'''
param(
  [Parameter(Mandatory=$true)][string]$TemplatePath,
  [Parameter(Mandatory=$true)][string]$CsvPath,
  [Parameter(Mandatory=$true)][string]$PdfPath
)
$ErrorActionPreference = "Stop"
$word = $null
$doc = $null
$merged = $null
try {
  $word = New-Object -ComObject Word.Application
  $word.Visible = $false
  $word.DisplayAlerts = 0
  foreach ($printerName in @("Microsoft Print to PDF", "OneNote (Desktop)")) {
    try {
      $word.ActivePrinter = $printerName
      break
    } catch {
    }
  }
  $doc = $word.Documents.Open($TemplatePath, $false, $true)
  $doc.MailMerge.GetType().InvokeMember('OpenDataSource', [System.Reflection.BindingFlags]::InvokeMethod, $null, $doc.MailMerge, @($CsvPath)) | Out-Null
  $doc.MailMerge.Destination = 0
  $doc.MailMerge.SuppressBlankLines = $true
  $doc.MailMerge.Execute($false)
  $merged = $word.ActiveDocument
  $merged.ExportAsFixedFormat($PdfPath, 17)
}
finally {
  if ($merged -ne $null) { $merged.Close($false) | Out-Null }
  if ($doc -ne $null) { $doc.Close($false) | Out-Null }
  if ($word -ne $null) { $word.Quit() | Out-Null }
}
'''
    with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as ps_file:
        ps_file.write(ps_script)
        ps_path = ps_file.name
    try:
        _check_cancel(cancel_event)
        startupinfo = None
        creationflags = 0
        if os.name == "nt":
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = 0
            creationflags = subprocess.CREATE_NO_WINDOW
        process = subprocess.Popen(
            [
                "powershell.exe",
                "-NoProfile",
                "-NonInteractive",
                "-WindowStyle",
                "Hidden",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                ps_path,
                "-TemplatePath",
                str(template_path),
                "-CsvPath",
                str(csv_path),
                "-PdfPath",
                str(pdf_path),
            ],
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
        with ACTIVE_PROCESS_LOCK:
            ACTIVE_PROCESSES.add(process)
        started_at = time.time()
        while process.poll() is None:
            if cancel_event is not None and cancel_event.is_set():
                _terminate_active_processes()
                raise ProcessingCancelled("Cancelled by user")
            if time.time() - started_at > 600:
                _terminate_active_processes()
                return False, "Timed out while Microsoft Word was generating this factory PDF."
            time.sleep(0.25)

        stdout, stderr = process.communicate()
        with ACTIVE_PROCESS_LOCK:
            ACTIVE_PROCESSES.discard(process)
        completed_returncode = process.returncode
        if completed_returncode != 0:
            message = (stderr or stdout or "Word mail merge failed.").strip()
            return False, message
        return True, None
    except ProcessingCancelled:
        raise
    finally:
        if "process" in locals():
            with ACTIVE_PROCESS_LOCK:
                ACTIVE_PROCESSES.discard(process)
        try:
            os.remove(ps_path)
        except OSError:
            pass


def _write_factory_label_pdfs(output_dir, shipment_df, cols, master_lookup, region, progress_callback=None, cancel_event=None):
    _check_cancel(cancel_event)
    template_path = _resolve_template(_label_template_name(region))
    if not template_path:
        return False, f"Label template '{_label_template_name(region)}' not found.", []

    region = str(region).strip().upper()
    if region == "TJXCA":
        order_lookup, dc_lookup = master_lookup
        label_rows = _build_tjxca_label_rows(shipment_df, cols, order_lookup, dc_lookup)
    else:
        label_rows, error = _build_label_rows(shipment_df, cols, master_lookup, region)
        if error:
            return False, error, []
    if not label_rows:
        return False, f"No {region} label rows were generated for PDF merge.", []

    rows_by_factory = _merge_rows_by_factory(label_rows, region)
    pdf_dir = output_dir / "Label PDFs"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    if region == "TJXCA":
        fieldnames = [
            "Ship_Merchandise_To",
            "TJX_PO",
            "Dept_No",
            "Vendor_Style",
            "Colour",
            "Total_Units",
            "Store_ready",
            "PreTicketed",
            "Inner",
            "Carton_",
            "Total_Carton_Qty",
            "COO",
            "Factory",
        ]
    else:
        fieldnames = [
            "TJX_PO",
            "Dept_No",
            "Vendor_Style_",
            "Description",
            "Total_units",
            "Carton_",
            "Total_Carton_Qty",
            "Country_",
            "Vendor_Packed",
            "PreTicketed",
            "M_25KG",
            "To",
            "Factory",
        ]

    generated_pdfs = []
    with tempfile.TemporaryDirectory(prefix="_label_merge_", dir=Path.cwd()) as temp_dir:
        temp_dir = Path(temp_dir)
        factory_items = sorted(rows_by_factory.items(), key=lambda item: item[0].lower())
        for index, (factory, rows) in enumerate(factory_items, start=1):
            _check_cancel(cancel_event)
            if progress_callback:
                progress_callback(
                    0.72 + (0.23 * (index - 1) / max(len(factory_items), 1)),
                    f"Generating label PDF {index}/{len(factory_items)}: {factory}"
                )
            safe_factory = _safe_path_token(factory, "Factory")
            csv_path = temp_dir / f"{safe_factory}.csv"
            temp_pdf_path = temp_dir / f"{safe_factory}.pdf"
            pdf_path = pdf_dir / f"{safe_factory} Labels.pdf"
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                writer.writeheader()
                for row in rows:
                    _check_cancel(cancel_event)
                    writer.writerow(row)

            success, message = _write_mail_merge_pdf(template_path, csv_path, temp_pdf_path, cancel_event=cancel_event)
            if not success:
                return False, f"Failed to create label PDF for factory '{factory}': {message}", generated_pdfs
            os.replace(temp_pdf_path, pdf_path)
            generated_pdfs.append(str(pdf_path))

    return True, None, generated_pdfs


def _format_number(value):
    try:
        number = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return _clean_text(value)
    if number == number.to_integral_value():
        return str(int(number))
    return format(number.normalize(), "f")


def _extract_indigo_po_items(input_file):
    input_file = Path(input_file)
    parser_errors = []
    pages_text = []
    try:
        import pypdfium2 as pdfium
        pdf = pdfium.PdfDocument(str(input_file))
        for page in pdf:
            textpage = page.get_textpage()
            pages_text.append(textpage.get_text_range() or "")
    except Exception as exc:
        parser_errors.append(f"pypdfium2: {exc}")

    text = "\n".join(pages_text)
    if not text.strip():
        details = "; ".join(parser_errors) if parser_errors else "no text extracted"
        return None, f"Unable to read Indigo PDF text ({details})."

    def _po_from_text(page_text, fallback=""):
        po_matches = re.findall(r"\b(4\d{9})\b", page_text)
        return po_matches[-1] if po_matches else fallback

    def _make_indigo_item(match, po):
        try:
            inner_packs = Decimal(match.group("inner").replace(",", ""))
            case_pack = Decimal(match.group("case").replace(",", ""))
            inner_qty = (case_pack / inner_packs).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except (InvalidOperation, ZeroDivisionError):
            return None, f"Invalid inner-pack/case-pack quantity near Indigo line {match.group('line')}."

        if inner_qty == inner_qty.to_integral_value():
            inner_qty_value = int(inner_qty)
        else:
            inner_qty_value = float(inner_qty)

        description = re.sub(r"\s+", " ", _clean_text(match.group("description")))
        sku = re.sub(r"\s+", "", _clean_text(match.group("sku")))
        if not sku:
            return None, f"Missing Manufacturer Part # near Indigo line {match.group('line')}."

        return {
            "po": po,
            "line": match.group("line"),
            "upc": match.group("upc"),
            "description": description,
            "order_qty": _format_number(match.group("qty")),
            "uom": _clean_text(match.groupdict().get("uom", "")),
            "case_pack_qty": _format_number(case_pack),
            "inner_packs": _format_number(inner_packs),
            "inner_qty": inner_qty_value,
            "sku": sku,
        }, None

    item_pattern = re.compile(
        r"(?P<line>\d{5})\s+"
        r"(?P<upc>\d{11,14})\s+"
        r"(?P<description>.*?)"
        r"#\s*of\s*Inner\s*Packs:\s*(?P<inner>[\d,]+(?:\.\d+)?).*?"
        r"Case\s*Pack\s*Qty:\s*(?P<case>[\d,]+(?:\.\d+)?).*?"
        r"Pack\s*UOM:\s*(?P<uom>[A-Za-z]+).*?"
        r"(?P<qty>[\d,]+(?:\.\d+)?)\s*[\r\n]+"
        r"Manufacturer's\s+Part\s*#:\s*(?P<sku>[A-Za-z0-9_ -]+?)(?=\s+Resale\s+Price:)",
        re.IGNORECASE | re.DOTALL,
    )

    items = []
    current_po = ""
    for page_text in pages_text:
        current_po = _po_from_text(page_text, current_po)
        for match in item_pattern.finditer(page_text):
            if not current_po:
                return None, f"Could not identify Indigo PO for line {match.group('line')}."
            item, error = _make_indigo_item(match, current_po)
            if error:
                return None, error
            items.append(item)

    if not items:
        return None, "No Indigo line items were found in the PDF."
    po_numbers = sorted({item["po"] for item in items})
    return {"po": po_numbers[0] if len(po_numbers) == 1 else "", "po_numbers": po_numbers, "items": items}, None


def _fit_text_lines(canvas_obj, text, max_width, font_name, font_size, max_lines=2):
    words = _clean_text(text).split()
    if not words:
        return [""]
    lines = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if canvas_obj.stringWidth(candidate, font_name, font_size) <= max_width or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
            if len(lines) >= max_lines:
                break
    if len(lines) < max_lines and current:
        lines.append(current)
    if words and len(lines) == max_lines:
        consumed = " ".join(lines).split()
        if len(consumed) < len(words):
            while lines[-1] and canvas_obj.stringWidth(lines[-1] + "...", font_name, font_size) > max_width:
                lines[-1] = " ".join(lines[-1].split()[:-1])
            lines[-1] = (lines[-1] + "...").strip()
    return lines


def _draw_indigo_inner_label(canvas_obj, item, x, y, width, height):
    from reportlab.graphics import renderPDF
    from reportlab.graphics.barcode import createBarcodeDrawing
    from reportlab.lib.units import inch

    pad_x = 0.12 * inch
    top = y + height - 0.16 * inch
    text_x = x + pad_x
    canvas_obj.setStrokeColorRGB(0.9, 0.9, 0.9)
    canvas_obj.setLineWidth(0.3)
    canvas_obj.rect(x + 1, y + 1, width - 2, height - 2, stroke=1, fill=0)

    canvas_obj.setFillColorRGB(0, 0, 0)
    canvas_obj.setFont("Helvetica-Bold", 9)
    canvas_obj.drawString(text_x, top, "SKU:")
    canvas_obj.setFont("Helvetica", 9)
    canvas_obj.drawString(text_x + 0.34 * inch, top, item["sku"])

    canvas_obj.setFont("Helvetica-Bold", 9)
    canvas_obj.drawString(text_x + 1.35 * inch, top, "UPC:")
    canvas_obj.setFont("Helvetica", 9)
    canvas_obj.drawString(text_x + 1.78 * inch, top, item["upc"])

    canvas_obj.setFont("Helvetica-Bold", 9)
    canvas_obj.drawString(text_x, top - 0.23 * inch, "Description:")
    canvas_obj.setFont("Helvetica", 8.5)
    desc_x = text_x + 0.86 * inch
    desc_width = width - (desc_x - x) - pad_x
    for idx, line in enumerate(_fit_text_lines(canvas_obj, item["description"], desc_width, "Helvetica", 8.5, max_lines=2)):
        canvas_obj.drawString(desc_x, top - 0.23 * inch - (idx * 0.15 * inch), line)

    canvas_obj.setFont("Helvetica-Bold", 10)
    canvas_obj.drawString(text_x, y + 0.23 * inch, "QTY:")
    canvas_obj.setFont("Helvetica-Bold", 13)
    canvas_obj.drawString(text_x + 0.42 * inch, y + 0.21 * inch, _format_number(item["inner_qty"]))

    barcode_value = re.sub(r"\D", "", item["upc"])
    barcode_type = "UPCA" if len(barcode_value) == 12 else "EAN13" if len(barcode_value) == 13 else "Code128"
    barcode = createBarcodeDrawing(
        barcode_type,
        value=barcode_value,
        barHeight=0.68 * inch,
        humanReadable=True,
    )
    max_barcode_width = width - (2 * pad_x)
    scale = min(max_barcode_width / barcode.width, 1.55)
    barcode_x = x + (width - (barcode.width * scale)) / 2
    barcode_y = y + 0.48 * inch
    canvas_obj.saveState()
    canvas_obj.translate(barcode_x, barcode_y)
    canvas_obj.scale(scale, scale)
    renderPDF.draw(barcode, canvas_obj, 0, 0)
    canvas_obj.restoreState()


def _write_indigo_sku_label_pdf(output_path, item, cancel_event=None):
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
    except ImportError:
        return False, "Missing PDF/barcode dependency: reportlab."

    _check_cancel(cancel_event)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    page_w, page_h = letter
    margin_x = 0.25 * inch
    margin_y = 0.25 * inch
    gap_x = 0.04 * inch
    cols = 2
    rows = 5
    label_w = (page_w - (2 * margin_x) - gap_x) / cols
    label_h = (page_h - (2 * margin_y)) / rows

    pdf = canvas.Canvas(str(output_path), pagesize=letter)
    pdf.setTitle(f"Indigo Inner Carton Labels - {item['sku']}")
    for row in range(rows):
        for col in range(cols):
            _check_cancel(cancel_event)
            x = margin_x + col * (label_w + gap_x)
            y = page_h - margin_y - ((row + 1) * label_h)
            _draw_indigo_inner_label(pdf, item, x, y, label_w, label_h)
    pdf.showPage()
    pdf.save()
    return True, None


def process_indigo_pdf_file(input_file, output_dir=None, progress_callback=None, cancel_event=None):
    _check_cancel(cancel_event)
    if progress_callback:
        progress_callback(0.10, "Reading Indigo PO PDF...")

    parsed, error = _extract_indigo_po_items(input_file)
    if error:
        return False, error, None

    input_file = Path(input_file)
    output_dir = Path(output_dir) if output_dir else Path.cwd()
    generated = []
    items = parsed["items"]
    for index, item in enumerate(items, start=1):
        _check_cancel(cancel_event)
        po = _safe_path_token(item["po"], "Indigo_PO")
        po_dir = output_dir / f"Indigo PO {po}"
        po_dir.mkdir(parents=True, exist_ok=True)
        if progress_callback:
            progress_callback(
                0.20 + (0.70 * (index - 1) / max(len(items), 1)),
                f"Generating Indigo inner labels {index}/{len(items)}: PO {item['po']} {item['sku']}"
            )
        safe_sku = _safe_path_token(item["sku"], f"SKU_{index}")
        output_path = po_dir / f"Indigo PO {po} {safe_sku} Inner Carton Labels.pdf"
        success, message = _write_indigo_sku_label_pdf(output_path, item, cancel_event=cancel_event)
        if not success:
            return False, message, None
        generated.append(output_path)

    if progress_callback:
        progress_callback(0.95, "Indigo labels complete.")
    return (
        True,
        f"Successfully generated {len(generated)} Indigo inner-carton label PDF(s) across {len(parsed.get('po_numbers', []))} PO folder(s).",
        str(output_dir),
    )


def _set_formula(ws, row, col, formula):
    ws.cell(row=row, column=col, value=formula)


def _write_aus_workbook(template_file, output_path, po, rows, keep_sheet=None):
    copy2(template_file, output_path)
    wb = load_workbook(output_path, keep_links=False)
    _strip_workbook_external_state(wb)

    actual_keep_sheet = keep_sheet
    if keep_sheet == "Packing List" and keep_sheet not in wb.sheetnames and "PL" in wb.sheetnames:
        actual_keep_sheet = "PL"

    if actual_keep_sheet and actual_keep_sheet not in wb.sheetnames:
        return False, f"AUS template must contain '{keep_sheet}' sheet."

    if "TJX Commercial Invoice" in wb.sheetnames:
        ws_ci = wb["TJX Commercial Invoice"]
        ws_ci["C6"].value = po
        ws_ci["I34"].value = datetime.now().date()
        ci_start = 13
        ci_rows = _build_ci_detail_rows(rows)
        ci_total = _find_row_by_value(ws_ci, "Total Units:", start_row=ci_start, col=1) or 24
        ci_total, ci_blank = _ensure_aus_ci_detail_capacity(ws_ci, ci_start, ci_total, len(ci_rows), 9)
        _copy_detail_rows_format(ws_ci, ci_blank, ci_start, ci_blank, 9)

        for idx, item in enumerate(ci_rows, start=ci_start):
            if item.get("row_type") == "composition":
                ws_ci.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=9)
                desc_cell = ws_ci.cell(idx, 5, item["desc"])
                desc_cell.alignment = copy_style(desc_cell.alignment)
                desc_cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    text_rotation=desc_cell.alignment.text_rotation,
                    wrap_text=False,
                    shrink_to_fit=desc_cell.alignment.shrink_to_fit,
                    indent=desc_cell.alignment.indent,
                )
                continue
            ws_ci.cell(idx, 2, item["sku"])
            ws_ci.cell(idx, 3, item["total_units"])
            ws_ci.cell(idx, 4, "CN")
            desc_cell = ws_ci.cell(idx, 5, item["desc"])
            desc_cell.alignment = copy_style(desc_cell.alignment)
            desc_cell.alignment = Alignment(
                horizontal=desc_cell.alignment.horizontal,
                vertical=desc_cell.alignment.vertical,
                text_rotation=desc_cell.alignment.text_rotation,
                wrap_text=False,
                shrink_to_fit=desc_cell.alignment.shrink_to_fit,
                indent=desc_cell.alignment.indent,
            )
            ws_ci.cell(idx, 8, item["price"])
            ws_ci.cell(idx, 9, item["amount"])

        last_ci = max(ci_start, ci_start + len(ci_rows) - 1)
        _set_formula(ws_ci, ci_total, 3, f"=SUM(C{ci_start}:C{last_ci})")
        _set_formula(ws_ci, ci_total, 9, f"=SUM(I{ci_start}:I{last_ci})")

    pl_sheet_name = "Packing List" if "Packing List" in wb.sheetnames else ("PL" if "PL" in wb.sheetnames else None)
    if pl_sheet_name:
        ws_pl = wb[pl_sheet_name]
        is_original_aus_pl = _clean_text(ws_pl["A1"].value).upper() == "PACKING LIST"
        if is_original_aus_pl:
            ws_pl["C14"].value = _as_excel_number(po)
            ws_pl["H13"].value = datetime.now().date()
            ws_pl["H14"].value = "China"
            pl_start = 16
            pl_total = _find_row_by_value(ws_pl, "TOTAL", start_row=pl_start, col=1) or 27
            pl_total = _ensure_detail_capacity(ws_pl, pl_start, pl_total, len(rows), 13, style_row=pl_start)
            _copy_detail_rows_format(ws_pl, pl_start, pl_start, pl_start + len(rows) - 1, 13)
            for idx, item in enumerate(rows, start=pl_start):
                ws_pl.cell(idx, 1, f'=IF(G{idx}=1,SUM($G${pl_start}:G{idx}),SUM($G${pl_start}:G{idx})-G{idx}+1 & "-" & SUM($G${pl_start}:G{idx}))')
                ws_pl.cell(idx, 2, item["sku"])
                ws_pl.cell(idx, 3, item["desc"])
                ws_pl.cell(idx, 4, "")
                ws_pl.cell(idx, 5, item["total_units"])
                ws_pl.cell(idx, 6, item["units"])
                ws_pl.cell(idx, 7, item["cartons"])
                ws_pl.cell(idx, 8, item["net_total"])
                ws_pl.cell(idx, 9, item["gross_total"])
                ws_pl.cell(idx, 10, "")
                ws_pl.cell(idx, 11, item["volume_total"])
                ws_pl.cell(idx, 12, _format_ctn_dimensions_cm(item["dimensions"]))
                ws_pl.cell(idx, 13, "")

            last_pl = max(pl_start, pl_start + len(rows) - 1)
            for col in [4, 5, 7, 8, 9, 10, 11]:
                _set_formula(ws_pl, pl_total, col, f"=SUM({ws_pl.cell(pl_start, col).column_letter}{pl_start}:{ws_pl.cell(last_pl, col).column_letter}{last_pl})")
            total_merge = f"A{pl_total}:C{pl_total}"
            if total_merge not in [str(merged_range) for merged_range in ws_pl.merged_cells.ranges]:
                ws_pl.merge_cells(total_merge)
        else:
            pl_start = 13
            pl_total = _find_row_by_value(ws_pl, "Totals", start_row=pl_start, col=1) or 28
            pl_total = _ensure_detail_capacity(ws_pl, pl_start, pl_total, len(rows), 13, style_row=pl_start)
            _copy_detail_rows_format(ws_pl, pl_start, pl_start, pl_start + len(rows) - 1, 13)
            for idx, item in enumerate(rows, start=pl_start):
                ws_pl.cell(idx, 1, item["ctn"])
                ws_pl.cell(idx, 2, item["sku"])
                ws_pl.cell(idx, 3, item["desc"])
                ws_pl.cell(idx, 5, item["total_units"])
                ws_pl.cell(idx, 6, item["units"])
                ws_pl.cell(idx, 7, item["cartons"])
                ws_pl.cell(idx, 8, item["net_each"])
                ws_pl.cell(idx, 9, item["net_total"])
                ws_pl.cell(idx, 10, item["gross_total"])
                ws_pl.cell(idx, 12, item["volume_total"])
                ws_pl.cell(idx, 13, _format_ctn_dimensions_cm(item["dimensions"]))

            last_pl = max(pl_start, pl_start + len(rows) - 1)
            for col in range(4, 13):
                _set_formula(ws_pl, pl_total, col, f"=SUM({ws_pl.cell(pl_start, col).column_letter}{pl_start}:{ws_pl.cell(last_pl, col).column_letter}{last_pl})")

    if actual_keep_sheet:
        if actual_keep_sheet in wb.sheetnames:
            wb[actual_keep_sheet].sheet_state = "visible"
        for sheet_name in ["TJX Commercial Invoice", "Packing List"]:
            if sheet_name != actual_keep_sheet and sheet_name in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb[sheet_name]
        if "PL" != actual_keep_sheet and "PL" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["PL"]
        if actual_keep_sheet in wb.sheetnames:
            wb.active = wb.sheetnames.index(actual_keep_sheet)

    if actual_keep_sheet != "PL":
        _force_workbook_arial_12(wb)
    wb.save(output_path)
    return True, None


def _write_eu_pl(template_file, output_path, po, rows):
    copy2(template_file, output_path)
    wb = load_workbook(output_path, keep_links=False)
    _strip_workbook_external_state(wb)
    ws = wb["PL"] if "PL" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws["C16"].value = po

    start = 20
    total_row = None
    for row in range(start, ws.max_row + 1):
        value = ws.cell(row, 6).value
        if isinstance(value, str) and value.startswith("=SUM("):
            total_row = row
            break
    total_row = total_row or 40
    total_row = _ensure_detail_capacity(ws, start, total_row, len(rows), 16, style_row=start)
    _copy_detail_rows_format(ws, start, start, start + len(rows) - 1, 16)

    for idx, item in enumerate(rows, start=start):
        ws.cell(idx, 1, item["ctn"])
        ws.cell(idx, 2, po)
        ws.cell(idx, 3, item["sku"])
        ws.cell(idx, 4, item["desc"])
        ws.cell(idx, 5, item["units"])
        ws.cell(idx, 6, item["cartons"])
        ws.cell(idx, 7, item["total_units"])
        ws.cell(idx, 8, item["net_each"])
        ws.cell(idx, 9, item["net_total"])
        ws.cell(idx, 10, item["gross_each"])
        ws.cell(idx, 11, item["gross_total"])
        ws.cell(idx, 12, _format_ctn_dimensions_cm(item["dimensions"]))
        ws.cell(idx, 13, item["volume_total"])
        ws.cell(idx, 14, item["teq_po"])

    last = max(start, start + len(rows) - 1)
    _set_formula(ws, total_row, 6, f"=SUM(F{start}:F{last})")
    _set_formula(ws, total_row, 7, f"=SUM(G{start}:G{last})")
    _set_formula(ws, total_row, 11, f"=SUM(K{start}:K{last})")
    _set_formula(ws, total_row, 13, f"=SUM(M{start}:M{last})")
    _force_workbook_arial_12(wb)
    wb.save(output_path)
    return True, None


def _write_eu_ci(template_file, output_path, po, rows):
    copy2(template_file, output_path)
    wb = load_workbook(output_path, keep_links=False)
    _strip_workbook_external_state(wb)
    ws = wb[wb.sheetnames[0]]
    ws["C13"].value = po
    ws["I14"].value = datetime.now().date()

    start = 18
    total_row = _find_row_by_value(ws, "Total", start_row=start, col=5) or 27
    total_row = _ensure_detail_capacity(ws, start, total_row, len(rows), 9, style_row=start)
    _copy_detail_rows_format(ws, start, start, start + len(rows) - 1, 9)
    for idx, item in enumerate(rows, start=start):
        ws.cell(idx, 1, item["ctn"])
        ws.cell(idx, 2, item["sku"])
        ws.cell(idx, 3, po)
        desc_cell = ws.cell(idx, 4, _ci_description(item))
        desc_cell.alignment = copy_style(desc_cell.alignment)
        desc_cell.alignment = Alignment(
            horizontal=desc_cell.alignment.horizontal,
            vertical=desc_cell.alignment.vertical,
            text_rotation=desc_cell.alignment.text_rotation,
            wrap_text=False,
            shrink_to_fit=desc_cell.alignment.shrink_to_fit,
            indent=desc_cell.alignment.indent,
        )
        ws.cell(idx, 5, item["units"])
        ws.cell(idx, 6, item["cartons"])
        ws.cell(idx, 7, item["total_units"])
        ws.cell(idx, 8, item["price"])
        ws.cell(idx, 9, item["amount"])

    last = max(start, start + len(rows) - 1)
    _set_formula(ws, total_row, 6, f"=SUM(F{start}:F{last})")
    _set_formula(ws, total_row, 7, f"=SUM(G{start}:G{last})")
    _set_formula(ws, total_row, 9, f"=SUM(I{start}:I{last})")
    _force_workbook_arial_12(wb)
    wb.save(output_path)
    return True, None


def _tjxca_split_address(destination):
    dc_name = _clean_text(destination.get("dc_name")) or "Winners Merchants International LP"
    ship_to = _clean_text(destination.get("ship_to"))
    parts = [part.strip() for part in ship_to.split(",") if part.strip()]
    if len(parts) >= 2:
        line1 = parts[0]
        line2 = ", ".join(parts[1:])
    else:
        line1 = ship_to
        line2 = ""
    return dc_name, line1, line2


def _write_tjxca_ci(template_file, output_path, po, rows, destination):
    copy2(template_file, output_path)
    wb = load_workbook(output_path, keep_links=False)
    _strip_workbook_external_state(wb)
    ws = wb[wb.sheetnames[0]]
    dc_name, line1, line2 = _tjxca_split_address(destination)
    ws["F5"].value = dc_name.upper()
    ws["H5"].value = dc_name.upper()
    ws["F6"].value = line1
    ws["H6"].value = line1
    ws["F7"].value = line2
    ws["H7"].value = line2
    ws["C13"].value = po
    ws["I13"].value = f"PO{po}"
    ws["I14"].value = datetime.now().date()

    start = 18
    total_row = _find_row_by_value(ws, "Total", start_row=start, col=5) or 37
    total_row = _ensure_detail_capacity(ws, start, total_row, len(rows), 9, style_row=start)
    _copy_detail_rows_format(ws, start, start, start + len(rows) - 1, 9)
    for idx, item in enumerate(rows, start=start):
        ws.cell(idx, 1, item["ctn"])
        ws.cell(idx, 2, item["sku"])
        ws.cell(idx, 3, po)
        ws.cell(idx, 4, item["desc"])
        ws.cell(idx, 5, item["units"])
        ws.cell(idx, 6, item["cartons"])
        ws.cell(idx, 7, item["total_units"])
        ws.cell(idx, 8, item["price"])
        ws.cell(idx, 9, item["amount"])

    last = max(start, start + len(rows) - 1)
    _set_formula(ws, total_row, 6, f"=SUM(F{start}:F{last})")
    _set_formula(ws, total_row, 7, f"=SUM(G{start}:G{last})")
    _set_formula(ws, total_row, 9, f"=SUM(I{start}:I{last})")
    _force_workbook_arial_12(wb)
    wb.save(output_path)
    return True, None


def _write_tjxca_pl(template_file, output_path, po, rows, destination):
    copy2(template_file, output_path)
    wb = load_workbook(output_path, keep_links=False)
    _strip_workbook_external_state(wb)
    ws = wb[wb.sheetnames[0]]
    dc_name, line1, line2 = _tjxca_split_address(destination)
    ws["J8"].value = dc_name.upper()
    ws["J9"].value = line1
    ws["J10"].value = line2
    ws["C16"].value = po

    start = 20
    total_row = _find_row_by_value(ws, "Total", start_row=start, col=5) or 39
    total_row = _ensure_detail_capacity(ws, start, total_row, len(rows), 12, style_row=start)
    _copy_detail_rows_format(ws, start, start, start + len(rows) - 1, 12)
    for idx, item in enumerate(rows, start=start):
        ws.cell(idx, 1, item["ctn"])
        ws.cell(idx, 2, item["sku"])
        ws.cell(idx, 3, _as_excel_number(po))
        ws.cell(idx, 4, item["desc"])
        ws.cell(idx, 5, item["units"])
        ws.cell(idx, 6, item["cartons"])
        ws.cell(idx, 7, item["total_units"])
        ws.cell(idx, 8, item["gross_each"])
        ws.cell(idx, 9, item["gross_total"])
        ws.cell(idx, 10, f'{_format_ctn_dimensions_cm(item["dimensions"])}cm' if item["dimensions"] else "")
        ws.cell(idx, 11, item["volume_total"])
        ws.cell(idx, 12, item["teq_po"])

    last = max(start, start + len(rows) - 1)
    _set_formula(ws, total_row, 6, f"=SUM(F{start}:F{last})")
    _set_formula(ws, total_row, 7, f"=SUM(G{start}:G{last})")
    _set_formula(ws, total_row, 11, f"=SUM(K{start}:K{last})")
    _force_workbook_arial_12(wb)
    wb.save(output_path)
    return True, None


def process_regional_summary_file(input_file, output_dir=None, region="AUS", progress_callback=None, cancel_event=None):
    _check_cancel(cancel_event)
    region = str(region).strip().upper()
    loaded, error = _load_region_summary(input_file, region)
    if error:
        return False, error, None

    df, cols = loaded
    master_lookup = _master_info_lookup(input_file)
    output_dir = Path(output_dir) if output_dir else Path.cwd()
    output_dir.mkdir(parents=True, exist_ok=True)
    input_stem = _safe_path_token(Path(input_file).stem, "Shipment")
    df["_ShipmentFolder"] = df.apply(lambda row: _shipment_key(row, cols, input_stem), axis=1)

    if region == "AUS":
        template_ci = _resolve_template("Template (AUS CI).xlsx")
        template_pl = _resolve_template("Template (AUS PL).xlsx")
        if not template_ci or not template_pl:
            return False, "Template files 'Template (AUS CI).xlsx' and/or 'Template (AUS PL).xlsx' not found.", None
    elif region == "EU":
        template_pl_50 = _resolve_template("Template (EU PL 50).xlsx")
        template_pl_55 = _resolve_template("Template (EU PL 55).xlsx")
        template_ci_50 = _resolve_template("Template (EU CI 50).xlsx")
        template_ci_55 = _resolve_template("Template (EU CI 55).xlsx")
        if not template_pl_50 or not template_pl_55 or not template_ci_50 or not template_ci_55:
            return False, "Template files 'Template (EU PL 50).xlsx', 'Template (EU PL 55).xlsx', 'Template (EU CI 50).xlsx', and/or 'Template (EU CI 55).xlsx' not found.", None
    else:
        return False, f"Region '{region}' is not implemented.", None

    generated_files = []
    label_files = []
    label_pdf_files = []
    shipment_dirs = []
    processed_pos = set()
    for shipment_name, shipment_df in df.groupby("_ShipmentFolder", sort=True):
        _check_cancel(cancel_event)
        if shipment_name == input_stem:
            shipment_dir = output_dir / shipment_name
        else:
            shipment_dir = output_dir / f"{region} {shipment_name}"
        ci_dir = shipment_dir / "CI"
        pl_dir = shipment_dir / "PL"
        ci_dir.mkdir(parents=True, exist_ok=True)
        pl_dir.mkdir(parents=True, exist_ok=True)
        shipment_dirs.append(str(shipment_dir))

        for po in sorted(shipment_df["_PONum_str"].dropna().unique(), key=str):
            _check_cancel(cancel_event)
            po_df = shipment_df[shipment_df["_PONum_str"] == po].copy()
            rows = _build_region_rows(po_df, cols, master_lookup)
            if not rows:
                continue
            processed_pos.add(po)

            if region == "AUS":
                ci_path = ci_dir / f"{po} CI.xlsx"
                pl_path = pl_dir / f"{po} PL.xlsx"
                success, message = _write_aus_workbook(
                    template_ci,
                    ci_path,
                    po,
                    rows,
                    keep_sheet="TJX Commercial Invoice",
                )
                if not success:
                    return False, message, None
                _check_cancel(cancel_event)
                success, message = _write_aus_workbook(
                    template_pl,
                    pl_path,
                    po,
                    rows,
                    keep_sheet="Packing List",
                )
                if not success:
                    return False, message, None
                generated_files.extend([str(ci_path), str(pl_path)])
            else:
                pl_path = pl_dir / f"TJX EU {po} PL.xlsx"
                ci_path = ci_dir / f"TJX EU {po} CI.xlsx"
                po_prefix = str(po)[:2]
                if po_prefix == "50":
                    template_pl = template_pl_50
                    template_ci = template_ci_50
                elif po_prefix == "55":
                    template_pl = template_pl_55
                    template_ci = template_ci_55
                else:
                    return False, f"EU PO '{po}' does not start with 50 or 55, so no CI/PL template can be selected.", None
                success, message = _write_eu_pl(template_pl, pl_path, po, rows)
                if not success:
                    return False, message, None
                _check_cancel(cancel_event)
                success, message = _write_eu_ci(template_ci, ci_path, po, rows)
                if not success:
                    return False, message, None
                generated_files.extend([str(pl_path), str(ci_path)])

        _check_cancel(cancel_event)
        labels_path = shipment_dir / "Labels.xlsx"
        success, message = _write_labels_workbook(labels_path, shipment_df, cols, master_lookup, region)
        if not success:
            return False, message, None
        label_files.append(str(labels_path))
        if progress_callback:
            progress_callback(0.70, "Generating label PDFs by factory...")
        success, message, pdfs = _write_factory_label_pdfs(
            shipment_dir,
            shipment_df,
            cols,
            master_lookup,
            region,
            progress_callback=progress_callback,
            cancel_event=cancel_event,
        )
        if not success:
            return False, message, None
        label_pdf_files.extend(pdfs)

    if not generated_files:
        return False, f"No {region} POs were exported.", None
    output_target = shipment_dirs[0] if len(shipment_dirs) == 1 else str(output_dir)
    label_message = f" and {len(label_files)} Labels workbook(s)" if label_files else ""
    pdf_message = f" and {len(label_pdf_files)} label PDF(s)" if label_pdf_files else ""
    return (
        True,
        f"Successfully processed {len(processed_pos)} {region} POs into {len(generated_files)} CI/PL files{label_message}{pdf_message}.",
        output_target,
    )


def process_tjxca_summary_file(input_file, output_dir=None, progress_callback=None, cancel_event=None):
    _check_cancel(cancel_event)
    loaded, error = _load_tjxca_summary(input_file)
    if error:
        return False, error, None

    df, cols = loaded
    master_lookup, order_lookup, dc_lookup = _tjxca_master_lookups(input_file)
    output_dir = Path(output_dir) if output_dir else Path.cwd()
    output_dir.mkdir(parents=True, exist_ok=True)
    input_stem = _safe_path_token(Path(input_file).stem, "TJX CA Shipment")
    df["_ShipmentFolder"] = df.apply(lambda row: _shipment_key(row, cols, input_stem), axis=1)

    template_ci = _resolve_template("Template (TJX CA CI).xlsx")
    template_pl = _resolve_template("Template (TJX CA PL).xlsx")
    if not template_ci or not template_pl:
        return False, "Template files 'Template (TJX CA CI).xlsx' and/or 'Template (TJX CA PL).xlsx' not found.", None
    if not _resolve_template(_label_template_name("TJXCA")):
        return False, "Label template 'TJX CA FOB Labels template.docx' not found.", None

    shipment_dirs = []
    generated_files = []
    label_files = []
    label_pdf_files = []
    processed_pos = set()
    for shipment_name, shipment_df in df.groupby("_ShipmentFolder", sort=True):
        _check_cancel(cancel_event)
        shipment_dir = output_dir / shipment_name
        ci_dir = shipment_dir / "CI"
        pl_dir = shipment_dir / "PL"
        ci_dir.mkdir(parents=True, exist_ok=True)
        pl_dir.mkdir(parents=True, exist_ok=True)
        shipment_dir.mkdir(parents=True, exist_ok=True)
        shipment_dirs.append(str(shipment_dir))

        for po in sorted(shipment_df["_PONum_str"].dropna().unique(), key=str):
            _check_cancel(cancel_event)
            po_df = shipment_df[shipment_df["_PONum_str"] == po].copy()
            rows = _build_region_rows(po_df, cols, master_lookup)
            if not rows:
                continue
            processed_pos.add(po)
            destination = _tjxca_destination_for_po(po, order_lookup, dc_lookup)
            ci_path = ci_dir / f"TJX CA {po} CI.xlsx"
            pl_path = pl_dir / f"TJX CA {po} Packing Slip.xlsx"
            success, message = _write_tjxca_ci(template_ci, ci_path, po, rows, destination)
            if not success:
                return False, message, None
            _check_cancel(cancel_event)
            success, message = _write_tjxca_pl(template_pl, pl_path, po, rows, destination)
            if not success:
                return False, message, None
            generated_files.extend([str(ci_path), str(pl_path)])

        labels_path = shipment_dir / "Labels.xlsx"
        if progress_callback:
            progress_callback(0.60, "Generating TJX CA Labels workbook...")
        _check_cancel(cancel_event)
        success, message = _write_tjxca_labels_workbook(labels_path, shipment_df, cols, order_lookup, dc_lookup)
        if not success:
            return False, message, None
        label_files.append(str(labels_path))

        if progress_callback:
            progress_callback(0.70, "Generating TJX CA label PDFs by factory...")
        success, message, pdfs = _write_factory_label_pdfs(
            shipment_dir,
            shipment_df,
            cols,
            (order_lookup, dc_lookup),
            "TJXCA",
            progress_callback=progress_callback,
            cancel_event=cancel_event,
        )
        if not success:
            return False, message, None
        label_pdf_files.extend(pdfs)

    output_target = shipment_dirs[0] if len(shipment_dirs) == 1 else str(output_dir)
    return (
        True,
        f"Successfully processed {len(processed_pos)} TJX CA PO(s) into {len(generated_files)} CI/PL files, {len(label_files)} Labels workbook(s), and {len(label_pdf_files)} label PDF(s).",
        output_target,
    )


def process_summary_file(input_file, output_dir=None, retailer="AUS", progress_callback=None, cancel_event=None):
    """
    Process the summary file and generate output.
    Returns (success, message, output_path)
    """
    try:
        INPUT = input_file
        retailer_key = _normalize_retailer_key(retailer)
        _check_cancel(cancel_event)
        if retailer_key in ("AUS", "EU"):
            return process_regional_summary_file(
                INPUT,
                output_dir,
                retailer_key,
                progress_callback=progress_callback,
                cancel_event=cancel_event,
            )
        if retailer_key == "TJXCA":
            return process_tjxca_summary_file(
                INPUT,
                output_dir,
                progress_callback=progress_callback,
                cancel_event=cancel_event,
            )
        if retailer_key == "INDIGO":
            return process_indigo_pdf_file(
                INPUT,
                output_dir,
                progress_callback=progress_callback,
                cancel_event=cancel_event,
            )
        return False, "Only TJX AUS, TJX EU, TJX CA summaries, and Indigo PO PDFs are supported in this tool.", None

        SHEET = "Packing Slip Data"

        # Load DataFrame
        try:
            df = pd.read_excel(INPUT, sheet_name=SHEET or 0)
        except Exception:
            xls = pd.ExcelFile(INPUT)
            df = pd.read_excel(INPUT, sheet_name=xls.sheet_names[0])

        # Column mapping for new summary format (Packing Slip Data sheet)
        COLS = {
            "PONum": "PONum",
            "SKU": "SKU",
            "ProductDescription": "Description",
            "Master_Qty": "Qty/CTN",
            "Carton_Qty": "Cartons",
            "Total_Units": "Grand Qty",
            "Factory": "Index",  # using Index column as placeholder
        }

        CONST = {
            "Dept#": "62",
            "Country of Origin": "CHINA",
            "Store Ready": "NO",
            "Pre-Ticketed": "NO",
        }

        # Safety: ensure columns exist
        required_cols = [
            COLS["PONum"],
            COLS["SKU"],
            COLS["ProductDescription"],
            COLS["Master_Qty"],
            COLS["Carton_Qty"],
        ]
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return False, f"Missing required column(s): {', '.join(missing)}", None

        # PO total cartons
        po_total_cartons = (
            df.groupby(COLS["PONum"])[COLS["Carton_Qty"]]
            .sum(min_count=1)
        )
        po_total_cartons = pd.to_numeric(po_total_cartons, errors="coerce").fillna(0).astype(int)

        df["_TotalCartonQty_byPO"] = df[COLS["PONum"]].map(po_total_cartons)

        # Check Total Carton Qty per PO is consistent
        check = df.groupby(COLS["PONum"])["_TotalCartonQty_byPO"].nunique()
        if not (check == 1).all():
            return False, "PO total carton quantity inconsistent!", None

        selected_pos = df[COLS["PONum"]].unique().tolist()
        df = df.copy()

        retailer_key = _normalize_retailer_key(retailer)

        # Validate retailer selection against summary content when possible.
        # We use "Orders from Fishbowl" -> CustomerName as the source signal.
        try:
            df_orders_check = pd.read_excel(INPUT, sheet_name="Orders from Fishbowl")
            if "CustomerName" in df_orders_check.columns:
                customer_names = df_orders_check["CustomerName"].dropna().astype(str).str.upper()
                looks_like_walmart = customer_names.str.contains("WALMART", regex=False).any()
                looks_like_tjx = customer_names.str.contains("TJX", regex=False).any()
                if retailer_key == "TJXCA" and looks_like_walmart:
                    return (
                        False,
                        "Selected retailer does not match this summary file. "
                        "File appears to be Walmart, but TJXCA is selected.",
                        None,
                    )
                if retailer_key == "WALMART" and not looks_like_walmart and looks_like_tjx:
                    return (
                        False,
                        "Selected retailer does not match this summary file. "
                        "File appears to be TJXCA, but Walmart is selected.",
                        None,
                    )
        except Exception:
            # If the sheet/column is unavailable, continue with normal logic.
            pass

        # Template (retailer-specific)
        if retailer_key == "TJXCA":
            template_name = "Template (TJXCA).xlsx"
        elif retailer_key == "WALMART":
            template_name = "Template (Walmart).xlsx"
        else:
            return False, f"Retailer '{retailer}' is not implemented yet.", None

        template_file = Path(template_name)
        if not template_file.exists():
            script_dir = Path(__file__).parent
            template_file = script_dir / template_name
            if not template_file.exists():
                return False, f"Template file '{template_name}' not found in current directory.", None

        output_dir = Path(output_dir) if output_dir else Path.cwd()

        # Walmart branch: generate one file per PO
        if retailer_key == "WALMART":
            def _normalize_po(v):
                s = str(v).strip()
                return s[:-2] if s.endswith(".0") else s

            # Read Fishbowl orders for WM item and ship-to
            orders_sheet = "Orders from Fishbowl"
            try:
                df_orders = pd.read_excel(INPUT, sheet_name=orders_sheet)
            except Exception:
                return False, f"Missing required sheet: '{orders_sheet}'", None

            po_col = "PONum"
            if po_col not in df_orders.columns:
                return False, "Orders from Fishbowl missing 'PONum' column.", None
            if "CustomerName" not in df_orders.columns or "ShipToAddress" not in df_orders.columns:
                return False, "Orders from Fishbowl missing 'CustomerName' or 'ShipToAddress' column.", None
            if "SKU" not in df_orders.columns:
                return False, "Orders from Fishbowl missing 'SKU' column.", None

            df_orders = df_orders.copy()
            df_orders["_PONum_str"] = df_orders[po_col].map(_normalize_po)
            df = df.copy()
            df["_PONum_str"] = df[COLS["PONum"]].map(_normalize_po)

            generated_files = []

            for po_raw in selected_pos:
                selected_po = _normalize_po(po_raw)
                po_df = df[df["_PONum_str"] == selected_po].copy()
                if po_df.empty:
                    continue

                orders_po = df_orders[df_orders["_PONum_str"] == selected_po]
                if orders_po.empty:
                    return False, f"No matching PO '{selected_po}' found in Orders from Fishbowl.", None

                if "ItemNote" in orders_po.columns:
                    wm_source = orders_po["ItemNote"]
                elif len(orders_po.columns) >= 14:
                    wm_source = orders_po.iloc[:, 13]  # Column N fallback
                else:
                    return False, "Orders from Fishbowl missing WM item source column (ItemNote / N).", None

                sku_to_wm = {}
                for sku, wm_item in zip(orders_po["SKU"], wm_source):
                    sku_to_wm[str(sku).strip()] = "" if pd.isna(wm_item) else str(wm_item).strip()

                customer_name = str(orders_po.iloc[0]["CustomerName"]).strip()
                ship_to_addr = str(orders_po.iloc[0]["ShipToAddress"]).strip()
                ship_to_value = f"{customer_name}\n{ship_to_addr}".strip()
                location_matches = re.findall(r"\d{8,}", customer_name)
                location_id = max(location_matches, key=len) if location_matches else ""
                customer_name_no_loc = customer_name
                if location_id:
                    customer_name_no_loc = customer_name_no_loc.replace(location_id, "")
                customer_name_no_loc = re.sub(r"\s*-\s*$", "", customer_name_no_loc).strip()
                ship_to_addr_one_line = re.sub(r"\s+", " ", ship_to_addr.replace("\n", ", ")).strip(" ,")
                if customer_name_no_loc and ship_to_addr_one_line:
                    ship_to_label = f"{customer_name_no_loc} - {ship_to_addr_one_line}"
                else:
                    ship_to_label = customer_name_no_loc or ship_to_addr_one_line

                output_path = output_dir / f"{selected_po}.xlsx"
                copy2(template_file, output_path)

                wb = load_workbook(output_path, keep_links=False)
                if wb.defined_names:
                    for name in list(wb.defined_names):
                        del wb.defined_names[name]
                for ws_rm in wb.worksheets:
                    if ws_rm.tables:
                        for table_name in list(ws_rm.tables.keys()):
                            del ws_rm.tables[table_name]
                if "PL" not in wb.sheetnames:
                    return False, "Template missing 'PL' sheet.", None

                ws = wb["PL"]
                ws["C12"].value = location_id
                ws["C13"].value = selected_po
                ws["G5"].value = ship_to_value

                # Clear old data from row 16 onward (columns A-I)
                start_row = 16
                start_col = 1
                end_col = 9
                max_clear_row = min(ws.max_row, start_row + 1000)

                merged_to_remove = [m for m in ws.merged_cells.ranges if m.min_row >= start_row]
                for m in merged_to_remove:
                    ws.unmerge_cells(str(m))

                for r in range(start_row, max_clear_row + 1):
                    for c in range(start_col, end_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if type(cell).__name__ != "MergedCell":
                            cell.value = None

                # Build order rows (not carton-expanded): one row per source SKU line
                order_rows = []
                carton_cursor = 1
                for _, row in po_df.iterrows():
                    sku = str(row[COLS["SKU"]]).strip()
                    wm_item = sku_to_wm.get(sku, "")
                    desc = row[COLS["ProductDescription"]]
                    units = int(pd.to_numeric(row[COLS["Master_Qty"]], errors="coerce") or 0)
                    cartons = int(pd.to_numeric(row[COLS["Carton_Qty"]], errors="coerce") or 0)
                    if cartons <= 0:
                        continue
                    end_carton = carton_cursor + cartons - 1
                    ctn_value = f"{carton_cursor}-{end_carton}" if cartons > 1 else str(carton_cursor)
                    order_rows.append([
                        ctn_value,            # A CTN #
                        sku,                 # B Stock #
                        wm_item,             # C WM item #
                        selected_po,         # D PO#
                        desc,                # E Description
                        units,               # F Qty/CTN
                        cartons,             # G Cartons
                        units * cartons      # H Grand Qty
                    ])
                    carton_cursor = end_carton + 1

                if not order_rows:
                    return False, f"No valid carton rows found for Walmart PO '{selected_po}'.", None

                rows_needed = len(order_rows)
                template_row_start = start_row
                template_row_end = 17
                template_rows = template_row_end - template_row_start + 1
                rows_to_add = rows_needed - template_rows

                if rows_to_add > 0:
                    insert_pos = template_row_end + 1
                    ws.insert_rows(insert_pos, amount=rows_to_add)

                    for i in range(rows_to_add):
                        src = ws[template_row_end]
                        dest = ws[template_row_end + 1 + i]
                        ws.row_dimensions[template_row_end + 1 + i].height = ws.row_dimensions[template_row_end].height
                        for cell_src, cell_dest in zip(src, dest):
                            if cell_src.has_style:
                                cell_dest._style = cell_src._style

                for i, row_data in enumerate(order_rows, start=start_row):
                    for j, val in enumerate(row_data, start=start_col):
                        ws.cell(row=i, column=j, value=val)

                last_data_row = start_row + len(order_rows) - 1
                total_row = last_data_row + 1
                ws.cell(row=total_row, column=6, value="Total")
                ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row}:G{last_data_row})")
                ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row}:H{last_data_row})")

                # Write Labels sheet (carton-expanded for Walmart)
                if "Labels" not in wb.sheetnames:
                    return False, "Template missing 'Labels' sheet.", None
                ws_labels = wb["Labels"]

                labels_start_row = 2
                max_clear_labels = min(ws_labels.max_row, labels_start_row + 1000)
                for r in range(labels_start_row, max_clear_labels + 1):
                    for c in range(1, 9):  # A-H
                        cell = ws_labels.cell(row=r, column=c)
                        if type(cell).__name__ != "MergedCell":
                            cell.value = None

                labels_rows = []
                carton_counter = 1
                total_cartons_po = int(
                    pd.to_numeric(po_df[COLS["Carton_Qty"]], errors="coerce").fillna(0).sum()
                )
                for _, row in po_df.iterrows():
                    sku = str(row[COLS["SKU"]]).strip()
                    wm_item = sku_to_wm.get(sku, "")
                    cartons = int(pd.to_numeric(row[COLS["Carton_Qty"]], errors="coerce") or 0)
                    if cartons <= 0:
                        continue
                    for _ in range(cartons):
                        labels_rows.append([
                            ship_to_label,     # A SHIP TO (without location number in customer name)
                            selected_po,       # B PO NUMBER
                            location_id,       # C LOCATION NUMBER
                            "33",              # D P.O. TYPE
                            "87",              # E DEPARTMENT
                            wm_item,           # F WAL-MART ITEM
                            carton_counter,    # G C (carton sequence)
                            total_cartons_po,  # H TC (total cartons)
                        ])
                        carton_counter += 1

                for i, row_data in enumerate(labels_rows, start=labels_start_row):
                    for j, val in enumerate(row_data, start=1):
                        ws_labels.cell(row=i, column=j, value=val)

                wb.save(output_path)
                generated_files.append(str(output_path))

            if not generated_files:
                return False, "No Walmart POs were exported.", None
            if len(generated_files) == 1:
                return True, "Successfully processed 1 PO.", generated_files[0]
            return True, f"Successfully processed {len(generated_files)} POs.", str(output_dir)

        # TJXCA branch: split output by PO prefix (e.g., 10 / 25)
        def get_tjx_profile(po_prefix):
            if po_prefix == "25":
                return {
                    "dept": "78",
                    "ship_to_address": "TJX Canada\n3185 American Drive\nMississauga, ON  L4V 1B8\nCANADA",
                }
            return {
                "dept": "62",
                "ship_to_address": "TJX Canada-WDC\n55 West Drive #891\nBrampton, ON  L6T 4A1\nCANADA",
            }

        po_groups = {}
        for po in selected_pos:
            po_short = extract_po_number(po)
            po_prefix = po_short[:2] if len(po_short) >= 2 else "10"
            po_groups.setdefault(po_prefix, []).append(po)

        ship_date_token = get_ship_date_filename_token(INPUT)
        is_mixed_prefix = len(po_groups) > 1
        generated_files = []

        for po_prefix, pos_in_group in po_groups.items():
            profile = get_tjx_profile(po_prefix)
            const_group = dict(CONST)
            const_group["Dept#"] = profile["dept"]

            df_group = df[df[COLS["PONum"]].isin(pos_in_group)].copy()
            df_group["CartonList"] = df_group[COLS["Carton_Qty"]].apply(
                lambda x: [1] * int(x) if pd.notnull(x) else []
            )
            df_group_expanded = df_group.explode("CartonList").reset_index(drop=True)
            total_cartons_group = len(df_group_expanded)
            df_group_expanded["Index.1"] = range(1, total_cartons_group + 1)
            df_group_expanded["_PO_short"] = df_group_expanded[COLS["PONum"]].apply(extract_po_number)

            po_numbers = [extract_po_number(p) for p in pos_in_group]
            po_text = "&".join(po_numbers)
            output_base = ship_date_token if ship_date_token else po_text
            output_stem = f"{output_base}({po_prefix})" if is_mixed_prefix else output_base
            output_path = output_dir / f"{output_stem}.xlsx"

            copy2(template_file, output_path)

            wb = load_workbook(output_path, keep_links=False)
            if wb.defined_names:
                for name in list(wb.defined_names):
                    del wb.defined_names[name]
            for ws in wb.worksheets:
                if ws.tables:
                    for table_name in list(ws.tables.keys()):
                        del ws.tables[table_name]
            if "PL" not in wb.sheetnames:
                return False, "Template missing 'PL' sheet.", None

            ws = wb["PL"]
            ws["C13"].value = "/".join(po_numbers)
            ws["G5"].value = profile["ship_to_address"]

            start_row = 16
            start_col = 1
            end_col = 9
            max_clear_row = start_row + 1000
            if ws.max_row < max_clear_row:
                max_clear_row = ws.max_row

            merged_to_remove = [m for m in ws.merged_cells.ranges if m.min_row >= start_row]
            for m in merged_to_remove:
                ws.unmerge_cells(str(m))

            for r in range(start_row, max_clear_row + 1):
                for c in range(start_col, end_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if type(cell).__name__ != "MergedCell":
                        cell.value = None

            order_rows = []
            carton_cursor = 1
            for idx, po in enumerate(pos_in_group):
                df_po = df_group[df_group[COLS["PONum"]] == po]
                po_short = extract_po_number(po)
                for _, row in df_po.iterrows():
                    sku = row[COLS["SKU"]]
                    desc = row[COLS["ProductDescription"]]
                    units = int(pd.to_numeric(row[COLS["Master_Qty"]], errors="coerce") or 0)
                    cartons = int(pd.to_numeric(row[COLS["Carton_Qty"]], errors="coerce") or 0)
                    if cartons <= 0:
                        continue
                    end_carton = carton_cursor + cartons - 1
                    ctn_range = f"{carton_cursor}-{end_carton}" if cartons > 1 else str(carton_cursor)
                    order_rows.append([
                        ctn_range,
                        sku,
                        "EX-work",
                        po_short,
                        desc,
                        units,
                        cartons,
                        units * cartons
                    ])
                    carton_cursor = end_carton + 1

                if idx < len(pos_in_group) - 1:
                    order_rows.append([""] * (end_col - start_col + 1))

            rows_needed = len(order_rows)
            template_row_start = start_row
            template_row_end = 17
            template_rows = template_row_end - template_row_start + 1
            rows_to_add = rows_needed - template_rows

            if rows_to_add > 0:
                insert_pos = template_row_end + 1
                ws.insert_rows(insert_pos, amount=rows_to_add)
                for i in range(rows_to_add):
                    src = ws[template_row_end]
                    dest = ws[template_row_end + 1 + i]
                    ws.row_dimensions[template_row_end + 1 + i].height = ws.row_dimensions[template_row_end].height
                    for cell_src, cell_dest in zip(src, dest):
                        if cell_src.has_style:
                            cell_dest._style = cell_src._style

            for i, row_data in enumerate(order_rows, start=start_row):
                for j, val in enumerate(row_data, start=start_col):
                    ws.cell(row=i, column=j, value=val)

            last_data_row = start_row + len(order_rows) - 1
            total_row = last_data_row + 1
            ws.cell(row=total_row, column=6, value="Total")
            ws.cell(row=total_row, column=7, value=f"=SUM(G{start_row}:G{last_data_row})")
            ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row}:H{last_data_row})")

            ws_labels = wb["Labels"]
            labels_start_row = 2
            max_clear_labels = ws_labels.max_row
            if max_clear_labels > 1000:
                max_clear_labels = labels_start_row + 1000
            for r in range(labels_start_row, max_clear_labels + 1):
                for c in range(1, 11):
                    cell = ws_labels.cell(row=r, column=c)
                    if type(cell).__name__ != "MergedCell":
                        cell.value = None

            labels_data = []
            for _, row in df_group_expanded.sort_values(by=[COLS["PONum"], "Index.1"]).iterrows():
                po_short = row["_PO_short"]
                sku = row[COLS["SKU"]]
                desc = row[COLS["ProductDescription"]]
                units = int(pd.to_numeric(row[COLS["Master_Qty"]], errors="coerce") or 0)
                carton_index = row["Index.1"]
                labels_data.append([
                    po_short,
                    const_group["Dept#"],
                    sku,
                    desc,
                    units,
                    carton_index,
                    total_cartons_group,
                    const_group["Country of Origin"],
                    const_group["Store Ready"],
                    const_group["Pre-Ticketed"]
                ])

            for i, row_data in enumerate(labels_data, start=labels_start_row):
                for j, val in enumerate(row_data, start=1):
                    ws_labels.cell(row=i, column=j, value=val)

            wb.save(output_path)
            generated_files.append(str(output_path))

        if not generated_files:
            return False, "No TJXCA POs were exported.", None
        if len(generated_files) == 1:
            return True, f"Successfully processed {len(selected_pos)} POs.", generated_files[0]
        return True, f"Successfully processed {len(selected_pos)} POs into {len(generated_files)} files.", str(output_dir)


    except Exception as e:
        return False, f"Error during processing:\n{str(e)}\n\n{traceback.format_exc()}", None


# -----------------------
# Helper: open folder
# -----------------------
def open_in_file_manager(path: str):
    try:
        p = Path(path)
        folder = p if p.is_dir() else p.parent
        if sys.platform.startswith("win"):
            os.startfile(str(folder))
        elif sys.platform == "darwin":
            subprocess.run(["open", str(folder)], check=False)
        else:
            subprocess.run(["xdg-open", str(folder)], check=False)
    except Exception:
        pass


# -----------------------
# Helper: open file with default application
# -----------------------
def open_file_with_default_app(path: str):
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", path], check=False)
        else:
            subprocess.run(["xdg-open", path], check=False)
    except Exception:
        pass


def load_saved_output_dir():
    try:
        settings_path = SETTINGS_FILE if SETTINGS_FILE.exists() else OLD_SETTINGS_FILE
        if settings_path.exists():
            data = json.loads(settings_path.read_text(encoding="utf-8"))
            output_dir = data.get("output_dir", "")
            if output_dir and Path(output_dir).exists():
                return output_dir
    except Exception:
        pass
    return str(Path.cwd())


def save_output_dir(path: str):
    try:
        SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
        SETTINGS_FILE.write_text(json.dumps({"output_dir": str(path)}, indent=2), encoding="utf-8")
    except Exception:
        pass


# -----------------------
# Helper: get icon path for frozen or script
# -----------------------
def get_icon_path():
    # List of possible locations
    possible_paths = []
    
    # If frozen, check sys._MEIPASS and sys.executable directory
    if getattr(sys, 'frozen', False):
        possible_paths.append(sys._MEIPASS)
        possible_paths.append(os.path.dirname(sys.executable))
    else:
        possible_paths.append(os.path.dirname(__file__))
    
    # Always check current working directory
    possible_paths.append(os.getcwd())
    
    # Deduplicate
    seen = set()
    for base in possible_paths:
        if not base or base in seen:
            continue
        seen.add(base)
        icon_path = os.path.join(base, "Logo.ico")
        if os.path.exists(icon_path):
            return icon_path
    
    # Final fallback: just "Logo.ico" in current directory (already checked)
    return None


# -----------------------
# Modal-style UI App (fixed size, no resize/maximize)
# -----------------------
class SummaryToPLApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NAME)
        # Set window icon
        icon_path = get_icon_path()
        if icon_path:
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass  # fallback silently
        self.center_window(560, 780)         # ✅ fixed size
        self.root.resizable(False, False)      # ✅ cannot resize or maximize
        self.root.configure(bg=BG_MAIN)

        self.file_path = None
        self.output_dir = load_saved_output_dir()
        self.retailer = ctk.StringVar(value="AUS")
        self._running = False
        self._cancel_event = threading.Event()
        self._worker_thread = None
        self._ui_queue = queue.Queue()
        self._last_output = None

        # Main container
        self.container = ctk.CTkFrame(self.root, fg_color=BG_MAIN, corner_radius=0)
        self.container.pack(fill="both", expand=True, padx=28, pady=14)

        # Title (changed)
        ctk.CTkLabel(
            self.container,
            text=APP_NAME,
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color=TEXT_DARK
        ).pack(anchor="center", pady=(4, 10))

        # Status
        self.status_badge = ctk.CTkLabel(
            self.container,
            text="● Idle",
            font=ctk.CTkFont(size=12, weight="bold"),
            text_color=TEXT_MUTED
        )
        self.status_badge.pack(anchor="w", pady=(0, 10))

        # Retailer selector (UI only for now)
        retailer_row = ctk.CTkFrame(self.container, fg_color="transparent")
        retailer_row.pack(fill="x", pady=(0, 8))

        ctk.CTkLabel(
            retailer_row,
            text="Option",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=TEXT_DARK
        ).pack(side="left")

        self.retailer_selector = ctk.CTkSegmentedButton(
            retailer_row,
            values=["AUS", "EU", "CA", "INDIGO"],
            variable=self.retailer,
            height=28,
            corner_radius=12,
            fg_color="#CBD5E1",
            selected_color="#3B82F6",
            selected_hover_color="#2563EB",
            unselected_color="#CBD5E1",
            unselected_hover_color="#CBD5E1",
            text_color="#FFFFFF",
            text_color_disabled="#94A3B8"
        )
        self.retailer_selector.pack(side="left", padx=(12, 0))

        # -----------------------
        # Dropzone Card
        # -----------------------
        self.drop_card = ctk.CTkFrame(
            self.container,
            fg_color=CARD_BG,
            corner_radius=16,
            border_width=1,
            border_color=BORDER_LIGHT,
            height=180  # Fixed height for better centering
        )
        self.drop_card.pack(fill="x", pady=(0, 10))
        self.drop_card.pack_propagate(False)  # Keep fixed height

        drop_inner = ctk.CTkFrame(self.drop_card, fg_color="transparent")
        drop_inner.pack(fill="both", expand=True, padx=22, pady=14)

        # Center everything both vertically and horizontally
        drop_inner.grid_rowconfigure(0, weight=1)
        drop_inner.grid_rowconfigure(1, weight=1)
        drop_inner.grid_rowconfigure(2, weight=1)
        drop_inner.grid_columnconfigure(0, weight=1)

        # Create a centered container for all widgets
        center_container = ctk.CTkFrame(drop_inner, fg_color="transparent")
        center_container.grid(row=1, column=0)  # Center row

        # Configure center_container to center its children - more top weight for better centering
        center_container.grid_rowconfigure(0, weight=2)  # More top space
        center_container.grid_rowconfigure(1, weight=0)  # Button
        center_container.grid_rowconfigure(2, weight=0)  # "or" label
        center_container.grid_rowconfigure(3, weight=0)  # Drag text
        center_container.grid_rowconfigure(4, weight=0)  # File name
        center_container.grid_rowconfigure(5, weight=3)  # More bottom space
        center_container.grid_columnconfigure(0, weight=1)

        # Browse button - centered
        self.btn_browse_file = ctk.CTkButton(
            center_container,
            text="⬆  Select a file",
            height=42,
            corner_radius=21,
            fg_color=CARD_BG,
            text_color=PRIMARY,
            hover_color=PRIMARY_HOVER,
            border_width=1,
            border_color=BORDER_LIGHT,
            command=self.select_file,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        self.btn_browse_file.grid(row=1, column=0, pady=(0, 8))

        # "or" label - centered
        or_label = ctk.CTkLabel(
            center_container,
            text="or",
            text_color=TEXT_MUTED,
            font=ctk.CTkFont(size=12)
        )
        or_label.grid(row=2, column=0, pady=(0, 8))

        # Drag and drop text - centered
        drag_label = ctk.CTkLabel(
            center_container,
            text="Drag and drop a file here",
            text_color="#374151",
            font=ctk.CTkFont(size=13)
        )
        drag_label.grid(row=3, column=0, pady=(0, 8))

        # File name label - centered
        self.file_name_label = ctk.CTkLabel(
            center_container,
            text="",
            text_color=SUCCESS,
            font=ctk.CTkFont(size=12, weight="bold"),
            wraplength=480,
            justify="center"
        )
        self.file_name_label.grid(row=4, column=0, pady=(8, 0))

        self.drop_card.drop_target_register(DND_FILES)
        self.drop_card.dnd_bind("<<Drop>>", self.on_file_drop)

        # -----------------------
        # Output Card
        # -----------------------
        out_card = ctk.CTkFrame(
            self.container,
            fg_color=CARD_BG,
            corner_radius=16,
            border_width=1,
            border_color=BORDER_LIGHT
        )
        out_card.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            out_card,
            text="Output directory",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=TEXT_DARK
        ).pack(anchor="w", padx=18, pady=(14, 6))

        self.output_path_label = ctk.CTkLabel(
            out_card,
            text=self.output_dir,
            text_color=SUCCESS,
            wraplength=500,
            justify="left"
        )
        self.output_path_label.pack(anchor="w", padx=18, pady=(0, 10))

        out_btn_row = ctk.CTkFrame(out_card, fg_color="transparent")
        out_btn_row.pack(fill="x", padx=18, pady=(0, 14))

        self.btn_browse_out = ctk.CTkButton(
            out_btn_row,
            text="Change",
            height=34,
            command=self.select_output_dir
        )
        self.btn_browse_out.pack(side="left", padx=(0, 10))

        self.btn_open_out = ctk.CTkButton(
            out_btn_row,
            text="Open Folder",
            height=34,
            command=self.open_output_folder
        )
        self.btn_open_out.pack(side="left")

        # -----------------------
        # Activity Card (now guaranteed visible)
        # -----------------------
        activity_card = ctk.CTkFrame(
            self.container,
            fg_color=CARD_BG,
            corner_radius=16,
            border_width=1,
            border_color=BORDER_LIGHT
        )
        activity_card.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            activity_card,
            text="Activity",
            font=ctk.CTkFont(size=13, weight="bold"),
            text_color=TEXT_DARK
        ).pack(anchor="w", padx=18, pady=(14, 8))

        self.log_box = ctk.CTkTextbox(activity_card, height=105, corner_radius=12)
        self.log_box.pack(fill="x", padx=18, pady=(0, 14))
        self.log_box.configure(state="disabled")

        # -----------------------
        # Progress
        # -----------------------
        self.progress = ctk.CTkProgressBar(self.container)
        self.progress.pack(fill="x", pady=(2, 4))
        self.progress.set(0)

        self.progress_label = ctk.CTkLabel(self.container, text="Ready", text_color=TEXT_MUTED)
        self.progress_label.pack(anchor="w", pady=(0, 10))

        # -----------------------
        # Actions
        # -----------------------
        actions = ctk.CTkFrame(self.container, fg_color="transparent")
        actions.pack(fill="x", pady=(0, 6))
        actions.grid_columnconfigure(0, weight=1)
        actions.grid_columnconfigure(1, weight=1)

        self.btn_start = ctk.CTkButton(
            actions,
            text="Start processing",
            height=44,
            font=ctk.CTkFont(size=15, weight="bold"),
            state="disabled",
            command=self.start_processing
        )
        self.btn_start.grid(row=0, column=0, padx=(0, 8), sticky="ew")

        self.btn_reset = ctk.CTkButton(
            actions,
            text="Reset",
            height=44,
            fg_color="#E5E7EB",
            text_color="#111827",
            hover_color="#D1D5DB",
            state="disabled",
            command=self.reset
        )
        self.btn_reset.grid(row=0, column=1, padx=(8, 0), sticky="ew")

        # Init
        self.log("App ready. Select AUS, EU, CA, or INDIGO. Indigo uses a PO PDF; the others use summary .xlsx files.")
        self.root.after(120, self._poll_ui_queue)
        
    def center_window(self, width, height):
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        x = (screen_w - width) // 2
        y = (screen_h - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    # -----------------------
    # UI helpers
    # -----------------------
    def log(self, msg: str):
        self.log_box.configure(state="normal")
        timestamp = time.strftime("%H:%M:%S")
        self.log_box.insert("end", f"[{timestamp}] {msg}\n")
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def set_status(self, label: str, color: str):
        self.status_badge.configure(text=label, text_color=color)

    def set_controls_enabled(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        self.btn_browse_file.configure(state=state)
        self.btn_browse_out.configure(state=state)
        self.btn_open_out.configure(state=state)
        self.retailer_selector.configure(state=state)

        if enabled:
            self.btn_start.configure(state=("normal" if self.file_path else "disabled"))
            self.btn_reset.configure(state=("normal" if self.file_path else "disabled"))
        else:
            self.btn_start.configure(state="disabled")
            self.btn_reset.configure(state="disabled")

    def _revert_reset_button(self):
        self.btn_reset.configure(
            text="Reset",
            fg_color="#E5E7EB",
            text_color="#111827",
            hover_color="#D1D5DB",
            command=self.reset,
            state="normal" if self.file_path else "disabled"
        )

    # -----------------------
    # File interactions
    # -----------------------
    def select_file(self):
        if _normalize_retailer_key(self.retailer.get()) == "INDIGO":
            filetypes = [("PDF Files", "*.pdf")]
            title = "Select Indigo PO PDF"
        else:
            filetypes = [("Excel Files", "*.xlsx")]
            title = "Select Summary File"
        path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        if path:
            self.set_file(path)

    def on_file_drop(self, event):
        file = event.data.strip("{}")
        retailer_key = _normalize_retailer_key(self.retailer.get())
        expected_ext = ".pdf" if retailer_key == "INDIGO" else ".xlsx"

        if " " in file and file.lower().count(expected_ext) > 1:
            parts = [p.strip("{}") for p in file.split() if p.lower().endswith(expected_ext)]
            file = parts[0] if parts else file

        if file.lower().endswith(expected_ext):
            self.set_file(file)
        else:
            if retailer_key == "INDIGO":
                messagebox.showerror("File Type Error", "Please drop an Indigo PO PDF file (*.pdf).")
            else:
                messagebox.showerror("File Type Error", "Please drop an Excel file (*.xlsx).")

    def set_file(self, path: str):
        self.file_path = path
        self.file_name_label.configure(text=os.path.basename(path))

        self.btn_start.configure(state="normal")
        self.btn_reset.configure(state="normal")
        self.progress.set(0)
        self.progress_label.configure(text="Ready to process")
        self.set_status("● Ready", SUCCESS)
        self.log(f"Selected file: {path}")

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="Select Output Directory", initialdir=self.output_dir)
        if dir_path:
            self.output_dir = dir_path
            save_output_dir(dir_path)
            self.output_path_label.configure(text=dir_path)
            self.log(f"Output directory set: {dir_path}")

    def open_output_folder(self):
        open_in_file_manager(self.output_dir)

    # -----------------------
    # Run / Reset / Cancel
    # -----------------------
    def reset(self):
        if self._running:
            return
        self.file_path = None
        self._last_output = None

        self.file_name_label.configure(text="")
        self.btn_start.configure(state="disabled")
        self.btn_reset.configure(state="disabled")
        self.progress.set(0)
        self.progress_label.configure(text="Ready")
        self.set_status("● Idle", TEXT_MUTED)
        self.log("Reset complete.")

    def cancel_processing(self):
        if not self._running:
            return
        self._cancel_event.set()
        _terminate_active_processes()
        self.progress_label.configure(text="Cancelling...")
        self.set_status("● Cancelling", WARNING)
        self.log("Cancellation requested. Stopping active work...")
        self.btn_reset.configure(state="disabled")

    # -----------------------
    # Processing (thread-safe)
    # -----------------------
    def start_processing(self):
        if self._running:
            return
        if not self.file_path:
            messagebox.showwarning("No File Selected", "Please select a file first.")
            return

        self._running = True
        self._processing_started_at = time.time()
        self._progress_label_base = "Starting..."
        self._progress_eta_deadline = None
        self._cancel_event.clear()
        save_output_dir(self.output_dir)
        self.set_status("● Processing", WARNING)
        self.set_controls_enabled(False)

        # Change Reset button to Cancel
        self.btn_reset.configure(
            text="Cancel",
            fg_color=ERROR,
            text_color="#FFFFFF",
            hover_color="#991B1B",
            command=self.cancel_processing,
            state="normal"
        )

        self.progress.set(0.12)
        self.progress_label.configure(text="Starting...")
        self.log("Processing started...")
        self.log(f"Option: {self.retailer.get()}")

        self._worker_thread = threading.Thread(target=self._worker, daemon=True)
        self._worker_thread.start()

    def _format_duration(self, seconds):
        seconds = max(0, int(seconds))
        if seconds < 60:
            return f"{seconds}s"
        minutes, seconds = divmod(seconds, 60)
        if minutes < 60:
            return f"{minutes}m {seconds}s"
        hours, minutes = divmod(minutes, 60)
        return f"{hours}h {minutes}m"

    def _refresh_progress_label(self):
        label = getattr(self, "_progress_label_base", "Ready")
        deadline = getattr(self, "_progress_eta_deadline", None)
        if self._running and deadline:
            remaining = max(0, int(deadline - time.time()))
            label = f"{label} | Time left {self._format_duration(remaining)}"
        self.progress_label.configure(text=label)

    def _set_progress_state(self, value, label):
        value = max(0.0, min(1.0, float(value)))
        self.progress.set(value)
        self._progress_label_base = label
        self._progress_eta_deadline = None
        started_at = getattr(self, "_processing_started_at", None)
        if self._running and started_at and 0.05 < value < 0.99:
            elapsed = time.time() - started_at
            if elapsed >= 10:
                remaining = max(0, int((elapsed / value) - elapsed))
                self._progress_eta_deadline = time.time() + remaining
        self._refresh_progress_label()

    def _worker(self):
        output_path = None
        retailer_key = _normalize_retailer_key(self.retailer.get())
        try:
            # Compute output path for potential deletion
            if retailer_key != "INDIGO":
                import pandas as pd
                from pathlib import Path
                df = pd.read_excel(self.file_path, sheet_name="Packing Slip Data")
                po_numbers = df["PONum"].apply(extract_po_number).unique()
                po_text = "&".join(po_numbers)
                output_dir = Path(self.output_dir)
                output_path = output_dir / f"{po_text}.xlsx"
        except Exception:
            # If we cannot compute, we'll still have output_path from process_summary_file later
            pass

        try:
            if retailer_key == "INDIGO":
                self._ui_queue.put(("progress", 0.25, "Reading Indigo PO PDF..."))
                next_status = "Generating Indigo inner carton labels..."
            else:
                self._ui_queue.put(("progress", 0.25, "Reading Excel..."))
                next_status = "Generating CI, PL, labels, and factory PDFs..."
            time.sleep(0.05)
            if self._cancel_event.is_set():
                raise Exception("Cancelled by user")

            self._ui_queue.put(("progress", 0.55, next_status))
            time.sleep(0.05)
            if self._cancel_event.is_set():
                raise Exception("Cancelled by user")

            def progress_callback(value, label):
                if self._cancel_event.is_set():
                    raise ProcessingCancelled("Cancelled by user")
                self._ui_queue.put(("progress", value, label))
                self._ui_queue.put(("log", label))

            success, message, output_path = process_summary_file(
                self.file_path,
                self.output_dir,
                retailer=self.retailer.get(),
                progress_callback=progress_callback,
                cancel_event=self._cancel_event,
            )

            if self._cancel_event.is_set():
                raise Exception("Cancelled by user")

            if success:
                self._last_output = output_path
                self._ui_queue.put(("progress", 1.0, "Done"))
                self._ui_queue.put(("log", f"{message} Saved: {output_path}"))
                self._ui_queue.put(("success", output_path))
            else:
                self._ui_queue.put(("progress", 0.0, "Failed"))
                self._ui_queue.put(("log", message))
                self._ui_queue.put(("error", message))

        except Exception as e:
            if isinstance(e, ProcessingCancelled) or "Cancelled by user" in str(e):
                _terminate_active_processes()
                # Delete unfinished file if exists
                if output_path and os.path.exists(output_path):
                    try:
                        os.remove(output_path)
                        self._ui_queue.put(("log", f"Cancelled. Deleted unfinished file: {output_path}"))
                    except Exception as rm_err:
                        self._ui_queue.put(("log", f"Failed to delete unfinished file: {rm_err}"))
                self._ui_queue.put(("cancelled",))
            else:
                err = f"Unexpected Error: {str(e)}\n{traceback.format_exc()}"
                self._ui_queue.put(("progress", 0.0, "Failed"))
                self._ui_queue.put(("log", err))
                self._ui_queue.put(("error", err))

        finally:
            self._ui_queue.put(("finish",))

    def _poll_ui_queue(self):
        try:
            while True:
                item = self._ui_queue.get_nowait()
                if not item:
                    continue

                kind = item[0]

                if kind == "log":
                    self.log(item[1])

                elif kind == "progress":
                    val, label = item[1], item[2]
                    self._set_progress_state(val, label)

                elif kind == "success":
                    out_path = item[1]
                    self.set_status("● Success", SUCCESS)
                    import tkinter.messagebox as mb
                    out_path_obj = Path(out_path)
                    is_folder = out_path_obj.exists() and out_path_obj.is_dir()
                    target = "folder" if is_folder else "file"
                    answer = mb.askyesno(
                        "Success",
                        f"Processing complete!\n\nSaved to:\n{out_path}\n\nDo you want to open the {target}?",
                        detail=f"Click 'Yes' to open the {target}, 'No' to close this dialog."
                    )
                    if answer:
                        if is_folder:
                            open_in_file_manager(out_path)
                        else:
                            open_file_with_default_app(out_path)
                elif kind == "error":
                    self.set_status("● Error", ERROR)
                    messagebox.showerror("Error", item[1])

                elif kind == "cancelled":
                    self.set_status("● Cancelled", TEXT_MUTED)
                    self.progress.set(0)
                    self._progress_eta_deadline = None
                    self.progress_label.configure(text="Cancelled")
                    self._revert_reset_button()

                elif kind == "finish":
                    self._running = False
                    self.set_controls_enabled(True)
                    self._progress_eta_deadline = None
                    self.progress_label.configure(text="Ready")
                    self._revert_reset_button()

        except queue.Empty:
            pass
        finally:
            if self._running:
                self._refresh_progress_label()
            self.root.after(120, self._poll_ui_queue)


def main():
    if len(sys.argv) >= 5 and sys.argv[1].lower() == "--process":
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
        region = sys.argv[2]
        input_file = sys.argv[3]
        output_dir = sys.argv[4]
        success, message, output_path = process_summary_file(input_file, output_dir, retailer=region)
        print(message)
        if output_path:
            print(output_path)
        sys.exit(0 if success else 1)

    root = TkinterDnD.Tk()
    app = SummaryToPLApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

